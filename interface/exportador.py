# interface/exportador.py (versão completa com melhorias visuais)
# ------------------------------------------------------------
# Exportadores Excel/PDF com formatação profissional e moderna
# Melhorias: cores sofisticadas, bordas elegantes, hierarquia
# visual aprimorada, formatação numérica consistente
#
# Autor: Valdinei Lankewicz
# Melhorias visuais: 2025
# ------------------------------------------------------------
from __future__ import annotations

import re
import unicodedata
import datetime as dt
from datetime import datetime
from tkinter import filedialog, messagebox
from difflib import get_close_matches

import pandas as pd
_pd = pd 
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Font, PatternFill, Border, Side, 
    numbers, GradientFill
)
from openpyxl.utils import get_column_letter as _gcl

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape, letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import cm, inch
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

# ============================================================
# DEFINIÇÃO DE ESTILOS PROFISSIONAIS
# ============================================================

# Paleta de cores profissional
CORES = {
    'primaria': '1F4788',      # Azul escuro profissional
    'secundaria': '4472C4',    # Azul médio
    'destaque': '2E75B6',      # Azul destaque
    'sucesso': '70AD47',       # Verde sucesso
    'alerta': 'FFC000',        # Amarelo/Laranja
    'erro': 'C00000',          # Vermelho
    'neutro_claro': 'F2F2F2',  # Cinza muito claro
    'neutro_medio': 'D9E1F2',  # Azul acinzentado claro
    'branco': 'FFFFFF',        # Branco
    'boletim': 'E7F0FF',       # Azul muito claro
    'ponto': 'E7FBEA',         # Verde muito claro
    'diferenca': 'FFF4E7',     # Laranja muito claro
}

def criar_borda(estilo='medium', cor='000000'):
    """Cria bordas padronizadas"""
    side = Side(style=estilo, color=cor)
    return Border(left=side, right=side, top=side, bottom=side)

# ------------------------------------------------------------
# Helper robusto para identificar linhas de TOTAL
# Cobre "TOTAL", "TOTAIS:", "TOTAL GERAL", etc.
# ------------------------------------------------------------
def _eh_total(valor) -> bool:
    s = str(valor or "").strip().upper()
    # aceita TOTAIS:, TOTAIS, TOTAL, TOTAL GERAL...
    return s.startswith("TOTAIS") or s.startswith("TOTAL")


def aplicar_cabecalho_profissional(ws, titulo, contrato, dt_ini, dt_fim, boletins, df_cols):
    """
    Aplica um cabeçalho no TOPO (linhas 1–4), sem usar ws.append,
    para não empurrar o conteúdo pro final quando o DF já foi escrito com startrow>=7.
    """
    max_col = max(1, len(df_cols))

    # === Linha 1: TÍTULO ===
    ws.row_dimensions[1].height = 30
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    c = ws.cell(row=1, column=1, value=str(titulo) if titulo is not None else "")
    c.font = Font(name='Calibri', size=16, bold=True, color=CORES['branco'])
    c.alignment = Alignment(horizontal="center", vertical="center")
    c.fill = PatternFill(start_color=CORES['primaria'], end_color=CORES['primaria'], fill_type="solid")

    # Helpers de estilo para rótulo (col A) e valor (col B..max_col)
    def _rotulo(r, texto):
        ws.row_dimensions[r].height = 22
        r_cell = ws.cell(row=r, column=1, value=texto)
        r_cell.font = Font(name='Calibri', size=11, bold=True, color=CORES['primaria'])
        r_cell.alignment = Alignment(horizontal="left", vertical="center")
        r_cell.fill = PatternFill(start_color=CORES['neutro_claro'], end_color=CORES['neutro_claro'], fill_type="solid")

    def _valor(r, texto):
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=max_col)
        v_cell = ws.cell(row=r, column=2, value=texto)
        v_cell.font = Font(name='Calibri', size=11)
        v_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

    # === Linha 2: CONTRATO ===
    _rotulo(2, "Contrato:")
    _valor(2, str(contrato) if contrato is not None else "")

    # === Linha 3: PERÍODO ===
    try:
        periodo_txt = f"{dt_ini:%d/%m/%Y} a {dt_fim:%d/%m/%Y}"
    except Exception:
        periodo_txt = ""
    _rotulo(3, "Período de Apuração:")
    _valor(3, periodo_txt)

    # === Linha 4: BOLETINS ===
    boletins_txt = "-" if (boletins is None or str(boletins).strip() == "") else str(boletins)
    _rotulo(4, "Boletins Incluídos:")
    _valor(4, boletins_txt)

    # Linhas 5–6 apenas como espaçamento visual (sem append)
    ws.row_dimensions[5].height = 6
    ws.row_dimensions[6].height = 6

def formatar_planilha_profissional(ws, startrow=5):
    """
    Aplica formatação profissional à planilha inteira
    startrow = linha onde começa o cabeçalho da tabela (padrão: 5)
    """
    max_row = ws.max_row
    max_col = ws.max_column
    
    # Ajusta largura das colunas
    for c in range(1, max_col + 1):
        col_letter = _gcl(c)
        if c == 1:  # Coluna de identificação (Nome, Data, etc)
            ws.column_dimensions[col_letter].width = 22
        else:
            ws.column_dimensions[col_letter].width = 15
    
    # Formata linha de cabeçalho da tabela
    ws.row_dimensions[startrow].height = 30
    for c in range(1, max_col + 1):
        cell = ws.cell(row=startrow, column=c)
        cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['branco'])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color=CORES['secundaria'], 
                                end_color=CORES['secundaria'], 
                                fill_type="solid")
        cell.border = criar_borda('medium', CORES['primaria'])
    
    # Formata linhas de dados
    for r in range(startrow + 1, max_row + 1):
        ws.row_dimensions[r].height = 18
        
        # Identifica se é linha de total
        primeira_celula = str(ws.cell(row=r, column=1).value or "")
        is_total = _eh_total(primeira_celula)
        
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            
            # Formatação para linha de total
            if is_total:
                cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['primaria'])
                cell.fill = PatternFill(start_color=CORES['neutro_medio'], 
                                       end_color=CORES['neutro_medio'], 
                                       fill_type="solid")
                cell.border = criar_borda('medium', CORES['primaria'])
                if c == 1:
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                # Formatação para linhas normais
                # Cor alternada para melhor leitura
                if (r - startrow) % 2 == 0:
                    cor_fundo = CORES['branco']
                else:
                    cor_fundo = CORES['neutro_claro']
                
                cell.fill = PatternFill(start_color=cor_fundo, 
                                       end_color=cor_fundo, 
                                       fill_type="solid")
                cell.border = criar_borda('thin', 'CCCCCC')
                
                if c == 1:
                    cell.font = Font(name='Calibri', size=9)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
                else:
                    cell.font = Font(name='Calibri', size=9)
                    cell.alignment = Alignment(horizontal="right", vertical="center")
                    
                    # Formata números
                    if isinstance(cell.value, (int, float)):
                        cell.number_format = '#,##0.00'
                        
                        # Destaca negativos em vermelho
                        if cell.value < 0:
                            cell.font = Font(name='Calibri', size=9, bold=True, color=CORES['erro'])
                    # Verifica strings numéricas negativas (formato HH:MM)
                    elif isinstance(cell.value, str) and cell.value.startswith('-'):
                        cell.font = Font(name='Calibri', size=9, bold=True, color=CORES['erro'])
    
    # Congela painéis no cabeçalho
    ws.freeze_panes = f"A{startrow + 1}"
    
    # Adiciona AutoFiltro
    if max_row > startrow:  # Só adiciona se houver dados
        ws.auto_filter.ref = f"A{startrow}:{_gcl(max_col)}{max_row}"

# ============================================================
# Integração com utils/formatadores (com fallbacks seguros)
# ============================================================
try:
    from utils.formatadores import (
        eh_negativo_str,
        somar_preservando_formato,
    )
    STARTROW_DADOS = 5  # Ajustado para novo layout compacto
except Exception:
    STARTROW_DADOS = 5  # Ajustado para novo layout compacto

    def eh_negativo_str(val) -> bool:
        if val is None:
            return False
        s = str(val).strip()
        if not s:
            return False
        if re.match(r"^-?\d{1,3}:\d{2}$", s):
            return s.startswith("-")
        s2 = s.replace(".", "").replace(",", ".")
        try:
            return float(s2) < 0
        except Exception:
            return s.startswith("-")

    def _is_hhmm(s: str) -> bool:
        return bool(re.match(r"^-?\d{1,3}:\d{2}$", str(s).strip()))

    def _to_minutes(s: str) -> int:
        m = re.match(r"^(-)?(\d{1,3}):(\d{2})$", str(s).strip())
        if not m:
            return 0
        sign = -1 if m.group(1) else 1
        return sign * (int(m.group(2)) * 60 + int(m.group(3)))

    def somar_preservando_formato(series) -> str:
        vals = [v for v in series if str(v).strip() not in ("", "-", "None", "nan")]
        if not vals:
            return "0:00"
        hhmm_count = sum(1 for v in vals if _is_hhmm(v))
        if hhmm_count >= max(1, len(vals) // 2):
            total_min = sum(_to_minutes(v) for v in vals if _is_hhmm(v))
            sign = "-" if total_min < 0 else ""
            total_min = abs(total_min)
            h, mi = divmod(total_min, 60)
            return f"{sign}{h}:{mi:02d}"
        acc = []
        for v in vals:
            s2 = str(v).replace(".", "").replace(",", ".")
            try:
                acc.append(float(s2))
            except Exception:
                return ""
        return f"{sum(acc):.2f}"

# ============================================================
# Constantes
# ============================================================
COLUNAS_PONTO = [
    "Data", "Nome", "CPF", "PIS",
    "Total Normais", "Total Noturno",
    "Extra 50%D", "Extra 100%D", "Extra 50%N", "Extra 100%N",
]

# ============================================================
# Helpers locais
# ============================================================

from utils.dataframe_utils import (
    resolver_base_ponto,
    preparar_df_ponto_para_comparacao,
    preparar_df_boletim_para_comparacao,
    groupby_sum_by_date,
)
from utils.comparacao import dfs_sem_ponto

def _montar_tripla_para_export(app, funcionario, data_ini, data_fim):
    """
    Retorna (df_b_f, df_p_f, df_d_f, sem_ponto) prontos para exportação.
    """
    # --- Boletim (B) ---
    df_b_base = app.dados_df[
        (app.dados_df["Funcionário"] == funcionario)
        & (app.dados_df["DATA"].dt.date >= data_ini)
        & (app.dados_df["DATA"].dt.date <= data_fim)
    ].copy()

    df_b_norm = preparar_df_boletim_para_comparacao(df_b_base)
    df_b = groupby_sum_by_date(
        df_b_norm.rename(columns={"DATA": "Data"}) if "DATA" in df_b_norm.columns else df_b_norm,
        "Data"
    )
    df_b_f = df_b.rename(columns={k: v for k, v in app.MAP_BOL.items() if k in df_b.columns}).copy()
    if "Data" in df_b_f.columns:
        df_b_f["Data"] = pd.to_datetime(df_b_f["Data"], errors="coerce").dt.floor("D")
    cols_b = ["Data"] + (["Boletim"] if "Boletim" in df_b_f.columns else []) + [h for h in app.HEADERS_VIZ if h in df_b_f.columns]
    df_b_f = df_b_f[cols_b] if "Data" in df_b_f.columns else pd.DataFrame(columns=["Data","Boletim"]+app.HEADERS_VIZ)

    # --- Ponto (P) ---
    df_ponto_sel, _, _ = resolver_base_ponto(
        app.df_ponto, app.df_relacao, funcionario, data_ini, data_fim, corte_similaridade=0.75
    )
    sem_ponto = (df_ponto_sel is None) or df_ponto_sel.empty

    if sem_ponto:
        df_p_f, df_d_f = dfs_sem_ponto(data_ini, data_fim, app.HEADERS_VIZ)
    else:
        df_p = preparar_df_ponto_para_comparacao(df_ponto_sel, data_ini, data_fim)
        inv_map_pto = {v: k for k, v in app.MAP_PTO.items()}
        df_p_f = df_p.rename(columns={k: inv_map_pto[k] for k in df_p.columns if k in inv_map_pto}).copy()
        if "Data" in df_p_f.columns:
            df_p_f["Data"] = pd.to_datetime(df_p_f["Data"], errors="coerce").dt.floor("D")
        cols_p = ["Data"] + [h for h in app.HEADERS_VIZ if h in df_p_f.columns]
        df_p_f = df_p_f[cols_p] if "Data" in df_p_f.columns else pd.DataFrame(columns=["Data"] + app.HEADERS_VIZ)

        # --- Diferença (D = B - P) ---
        df_m = pd.merge(df_b_f, df_p_f, on="Data", how="outer", suffixes=("_B", "_P")).sort_values("Data")
        for h in app.HEADERS_VIZ:
            if f"{h}_B" not in df_m.columns: df_m[f"{h}_B"] = 0.0
            if f"{h}_P" not in df_m.columns: df_m[f"{h}_P"] = 0.0
            df_m[f"{h}_B"] = pd.to_numeric(df_m[f"{h}_B"], errors="coerce").fillna(0.0)
            df_m[f"{h}_P"] = pd.to_numeric(df_m[f"{h}_P"], errors="coerce").fillna(0.0)
            df_m[h] = df_m[f"{h}_B"] - df_m[f"{h}_P"]
        df_d_f = df_m[["Data"] + app.HEADERS_VIZ].copy()

    return df_b_f, df_p_f, df_d_f, sem_ponto

def _get_headers_from_sheet(sheet):
    if sheet is None:
        return None
    h = getattr(sheet, "headers", None)
    return list(h() if callable(h) else h) if h else None

def _df_from_sheet_safe(sheet, expected_cols):
    data = [row[:] for row in sheet.get_sheet_data()] if sheet else []
    if not data:
        return pd.DataFrame(columns=expected_cols)
    row_len = max(len(r) for r in data)
    data = [(r + [""] * (row_len - len(r)))[:row_len] for r in data]
    cols = list(expected_cols[:row_len])
    if len(cols) < row_len:
        cols += [f"Col{idx}" for idx in range(len(cols) + 1, row_len + 1)]
    return pd.DataFrame(data, columns=cols)

def _sheet_name_unique(base, used: set) -> str:
    name = re.sub(r"[][:\\/?*]+", "", str(base)).strip() or "Aba"
    name = name[:31]
    orig = name
    i = 2
    while name in used:
        suf = f"_{i}"
        name = orig[: 31 - len(suf)] + suf
        i += 1
    used.add(name)
    return name

def _ult5_contratos(s):
    if s is None:
        return ""
    itens = [p.strip() for p in str(s).split(",") if p.strip()]
    return ", ".join([(i[-5:] if len(i) >= 5 else i) for i in itens])

def _canon_nome(nome: str) -> str:
    if not isinstance(nome, str):
        nome = "" if nome is None else str(nome)
    nome_norm = unicodedata.normalize("NFKC", nome)
    nome_norm = " ".join(nome_norm.strip().split())
    return nome_norm.upper()

def _safe_filename(name: str) -> str:
    name = str(name).strip()
    name = re.sub(r'[\\/:*?"<>|]+', "_", name)
    return name[:120] or "Contrato"

# ============================================================
# Exportações simples (Aba 1 e 2)
# ============================================================

def exportar_para_excel(app):
    """Exportação simples com formatação profissional"""
    if app.df_filtrado is None or app.df_filtrado.empty:
        messagebox.showwarning("Aviso", "Não há dados para exportar.")
        return
    
    arquivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx")],
        title="Salvar Relatório Excel Como",
        initialfile=f"Relatorio_{app.combo_funcionario.get()}.xlsx",
    )
    if not arquivo:
        return

    df = app.df_filtrado[app.colunas_visiveis].copy()
    if "DATA" in df.columns:
        df["DATA"] = pd.to_datetime(df["DATA"], errors="coerce").dt.strftime("%d/%m/%Y")

    # Calcula totais
    somas = {}
    for col in getattr(app, "colunas_para_somar", []):
        try:
            somas[col] = pd.to_numeric(df[col], errors="coerce").fillna(0).sum()
        except Exception:
            pass
    
    if somas:
        total = pd.Series({**{c: "" for c in df.columns}, **somas}, name="TOTAL")
        if "DATA" in total.index:
            total["DATA"] = "TOTAL"
        df = pd.concat([df, total.to_frame().T], ignore_index=True)

    # Cria workbook com formatação
    wb = Workbook()
    ws = wb.active
    ws.title = "Relatório"
    
    # Escreve cabeçalho
    for col_idx, col_name in enumerate(df.columns, start=1):
        ws.cell(row=STARTROW_DADOS, column=col_idx, value=col_name)
    
    # Escreve dados
    for row_idx, (_, row) in enumerate(df.iterrows(), start=STARTROW_DADOS + 1):
        for col_idx, col_name in enumerate(df.columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[col_name])
    
    # Aplica formatação profissional
    funcionario = app.combo_funcionario.get()
    periodo_ini = app.cal_data_inicial.entry.get()
    periodo_fim = app.cal_data_final.entry.get()
    
    try:
        dt_ini = datetime.strptime(periodo_ini, "%d/%m/%Y")
        dt_fim = datetime.strptime(periodo_fim, "%d/%m/%Y")
    except:
        dt_ini = datetime.now()
        dt_fim = datetime.now()
    
    aplicar_cabecalho_profissional(
        ws,
        titulo=f"Relatório de Horas - {funcionario}",
        contrato="-",
        dt_ini=dt_ini,
        dt_fim=dt_fim,
        boletins="-",
        df_cols=list(df.columns)
    )
    
    formatar_planilha_profissional(ws, startrow=STARTROW_DADOS)
    
    try:
        wb.save(arquivo)
        messagebox.showinfo("Sucesso", f"Relatório exportado para\n'{arquivo}'")
    except Exception as e:
        messagebox.showerror("Erro de Exportação", str(e))

def exportar_para_pdf(app):
    if app.df_filtrado is None or app.df_filtrado.empty:
        messagebox.showwarning("Aviso", "Não há dados para exportar.")
        return
    arquivo = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("Arquivos PDF", "*.pdf")],
        title="Salvar Relatório PDF Como",
        initialfile=f"Relatorio_{app.combo_funcionario.get()}.pdf",
    )
    if not arquivo:
        return

    df = app.df_filtrado[app.colunas_visiveis].copy()
    if "DATA" in df.columns:
        df["DATA"] = pd.to_datetime(df["DATA"], errors="coerce").dt.strftime("%d/%m/%Y")

    dados = [list(df.columns)] + df.values.tolist()

    somas = {}
    for col in getattr(app, "colunas_para_somar", []):
        try:
            somas[col] = pd.to_numeric(app.df_filtrado[col], errors="coerce").fillna(0).sum()
        except Exception:
            pass
    if somas:
        linha_total = ["TOTAL" if c == "DATA" else f"{somas.get(c, ''):.2f}" if c in somas else "" for c in df.columns]
        dados.append(linha_total)

    doc = SimpleDocTemplate(arquivo, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    story = [
        Paragraph("Relatório de Horas", styles["Title"]),
        Paragraph(f"Funcionário: {app.combo_funcionario.get()}", styles["Normal"]),
        Paragraph(
            f"Período: {app.cal_data_inicial.entry.get()} a {app.cal_data_final.entry.get()}",
            styles["Normal"],
        ),
        Spacer(1, 0.2 * inch),
    ]

    tabela = Table(dados)
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -2), colors.HexColor("#B4C6E7")),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#8EA9DB")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )
    story.append(tabela)
    try:
        doc.build(story)
        messagebox.showinfo("Sucesso", f"Relatório exportado para\n'{arquivo}'")
    except Exception as e:
        messagebox.showerror("Erro de Exportação", str(e))

# ============================================================
# Exportações de Ponto (Aba 2)
# ============================================================

def exportar_ponto_para_excel(app, completo: bool = False):
    df = app.df_ponto if completo else app.df_ponto_filtrado
    if df is None or df.empty:
        messagebox.showwarning("Aviso", "Nenhum dado de ponto disponível para exportar.")
        return

    if completo:
        nome_arquivo = "RegistroPonto_Completo.xlsx"
    else:
        funcionario = app.combo_funcionario.get().strip().replace(" ", "_")
        data_ini = dt.datetime.strptime(app.cal_data_inicial.entry.get(), "%d/%m/%Y")
        nome_arquivo = f"RegistroPonto_{funcionario}_{data_ini.strftime('%Y%m')}.xlsx"

    arquivo = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivo Excel", "*.xlsx")],
        initialfile=nome_arquivo,
        title="Salvar Relatório de Ponto",
    )
    if not arquivo:
        return

    df_export = df.copy()
    df_export["Data"] = pd.to_datetime(df_export["Data"]).dt.strftime("%d/%m/%Y")
    df_export = df_export[COLUNAS_PONTO]

    soma_linha = {
        col: pd.to_numeric(df_export[col], errors="coerce").fillna(0).sum()
        for col in df_export.columns
        if col not in ["Data", "Nome", "CPF", "PIS"]
    }
    total_row = pd.DataFrame(
        [
            {
                "Data": "TOTAIS:",
                "Nome": "",
                "CPF": "",
                "PIS": "",
                **{k: round(v, 2) for k, v in soma_linha.items()},
            }
        ]
    )
    df_export = pd.concat([df_export, total_row], ignore_index=True)

    # Cria workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Registro de Ponto"
    
    # Escreve dados
    for col_idx, col_name in enumerate(df_export.columns, start=1):
        ws.cell(row=STARTROW_DADOS, column=col_idx, value=col_name)
    
    for row_idx, (_, row) in enumerate(df_export.iterrows(), start=STARTROW_DADOS + 1):
        for col_idx, col_name in enumerate(df_export.columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=row[col_name])
    
    # Aplica formatação
    try:
        dt_ini = dt.datetime.strptime(app.cal_data_inicial.entry.get(), "%d/%m/%Y")
        dt_fim = dt.datetime.strptime(app.cal_data_final.entry.get(), "%d/%m/%Y")
    except:
        dt_ini = dt.datetime.now()
        dt_fim = dt.datetime.now()
    
    titulo = "Registro de Ponto (Completo)" if completo else f"Registro de Ponto - {app.combo_funcionario.get()}"
    
    aplicar_cabecalho_profissional(
        ws,
        titulo=titulo,
        contrato="-",
        dt_ini=dt_ini,
        dt_fim=dt_fim,
        boletins="-",
        df_cols=list(df_export.columns)
    )
    
    formatar_planilha_profissional(ws, startrow=STARTROW_DADOS)
    
    try:
        wb.save(arquivo)
        messagebox.showinfo("Sucesso", f"Relatório exportado para\n'{arquivo}'")
    except Exception as e:
        messagebox.showerror("Erro de Exportação", str(e))

def exportar_ponto_para_pdf(app, completo: bool = False):
    df = app.df_ponto if completo else app.df_ponto_filtrado
    if df is None or df.empty:
        messagebox.showwarning("Aviso", "Nenhum dado de ponto disponível para exportar.")
        return

    if completo:
        nome_arquivo = "RegistroPonto_Completo.pdf"
    else:
        funcionario = app.combo_funcionario.get().strip().replace(" ", "_")
        data_ini = dt.datetime.strptime(app.cal_data_inicial.entry.get(), "%d/%m/%Y")
        nome_arquivo = f"RegistroPonto_{funcionario}_{data_ini.strftime('%Y%m')}.pdf"

    arquivo = filedialog.asksaveasfilename(
        defaultextension=".pdf",
        filetypes=[("Arquivo PDF", "*.pdf")],
        initialfile=nome_arquivo,
        title="Salvar Relatório de Ponto",
    )
    if not arquivo:
        return

    doc = SimpleDocTemplate(arquivo, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    story = []

    titulo = "Relatório de Ponto (Completo)" if completo else "Relatório de Ponto do Funcionário"
    story.append(Paragraph(titulo, styles["Title"]))
    if not completo:
        funcionario = app.combo_funcionario.get()
        periodo = f"{app.cal_data_inicial.entry.get()} a {app.cal_data_final.entry.get()}"
        story.append(Paragraph(f"Funcionário: {funcionario}", styles["Normal"]))
        story.append(Paragraph(f"Período: {periodo}", styles["Normal"]))
    story.append(Spacer(1, 12))

    df_export = df.copy()
    df_export["Data"] = pd.to_datetime(df_export["Data"]).dt.strftime("%d/%m/%Y")
    df_export = df_export[COLUNAS_PONTO]

    dados = [COLUNAS_PONTO] + df_export.values.tolist()
    soma_linha = ["TOTAIS:", "", "", ""] + [
        f"{pd.to_numeric(df_export[col], errors='coerce').fillna(0).sum():.2f}"
        for col in COLUNAS_PONTO[4:]
    ]
    dados.append(soma_linha)

    tabela = Table(dados)
    tabela.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4472C4")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("BACKGROUND", (0, 1), (-1, -2), colors.HexColor("#D9E1F2")),
                ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#B4C6E7")),
                ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
            ]
        )
    )
    try:
        doc.build(story + [tabela])
        messagebox.showinfo("Sucesso", f"Relatório exportado para\n'{arquivo}'")
    except Exception as e:
        messagebox.showerror("Erro de Exportação", str(e))

# ============================================================
# Comparação por funcionário (Aba 4)
# ============================================================

def _gerar_dados_comparacao_funcionario(app, funcionario, data_ini_dt, data_fim_dt, contrato_id_especifico=None):
    """Retorna (df_boletim, df_ponto, df_diferenca, registro, boletins_set)."""
    df_boletim_orig = app.dados_df.copy()
    df_ponto_orig = app.df_ponto.copy()

    df_boletim_orig["DATA"] = pd.to_datetime(df_boletim_orig["DATA"], errors="coerce")
    df_ponto_orig["Data"] = pd.to_datetime(df_ponto_orig["Data"], errors="coerce")

    # Boletim base
    filtro_base = (
        (df_boletim_orig["Funcionário"].str.upper() == funcionario.upper())
        & (df_boletim_orig["DATA"].dt.date >= data_ini_dt.date())
        & (df_boletim_orig["DATA"].dt.date <= data_fim_dt.date())
    )
    df_boletim_filt = df_boletim_orig[filtro_base].copy()

    dias_do_contrato = None
    if contrato_id_especifico and not df_boletim_filt.empty:
        df_boletim_filt = df_boletim_filt[df_boletim_filt["Contrato"] == contrato_id_especifico].copy()
        dias_do_contrato = df_boletim_filt["DATA"].dt.date.unique()

    # Ponto: escolha de nome similar
    nomes_ponto = df_ponto_orig["Nome"].dropna().unique()
    nome_similar = get_close_matches(funcionario.upper(), [n.upper() for n in nomes_ponto], n=1, cutoff=0.75)
    df_ponto_filt = pd.DataFrame()
    if nome_similar:
        nome_ponto = next((n for n in nomes_ponto if n.upper() == nome_similar[0]), None)
        if nome_ponto:
            df_ponto_base = df_ponto_orig[
                (df_ponto_orig["Nome"].str.upper() == nome_ponto.upper())
                & (df_ponto_orig["Data"].dt.date >= data_ini_dt.date())
                & (df_ponto_orig["Data"].dt.date <= data_fim_dt.date())
            ].copy()
            df_ponto_filt = (
                df_ponto_base if dias_do_contrato is None else df_ponto_base[df_ponto_base["Data"].dt.date.isin(dias_do_contrato)].copy()
            )

    campos = [
        ("HORA NORMAL", "Total Normais"),
        ("H.N.", "Total Noturno"),
        ("H.E.", "Extra 50%D"),
        ("H.E.D.", "Extra 100%D"),
        ("H.E.N.", "Extra 50%N"),
        ("H.E.N.D.", "Extra 100%N"),
    ]
    colunas_resultado = app.headers_comparacao

    bgrp = (df_boletim_filt.groupby(df_boletim_filt["DATA"].dt.date).sum(numeric_only=True)) if not df_boletim_filt.empty else pd.DataFrame()
    pgrp = (df_ponto_filt.groupby(df_ponto_filt["Data"].dt.date).sum(numeric_only=True)) if not df_ponto_filt.empty else pd.DataFrame()

    datas = sorted(set(bgrp.index).union(pgrp.index))
    dados_boletim, dados_ponto, dados_dif = [], [], []

    for data in datas:
        lb = [data.strftime("%d/%m/%Y")]
        lp = [data.strftime("%d/%m/%Y")]
        ld = [data.strftime("%d/%m/%Y")]
        for campo_b, campo_p in campos:
            vb = bgrp.at[data, campo_b] if (data in bgrp.index and campo_b in bgrp.columns) else 0
            vp = pgrp.at[data, campo_p] if (data in pgrp.index and campo_p in pgrp.columns) else 0
            d = vb - vp
            lb.append(f"{vb:.2f}" if vb else "-")
            lp.append(f"{vp:.2f}" if vp else "-")
            ld.append(f"{d:.2f}" if d else "-")
        dados_boletim.append(lb)
        dados_ponto.append(lp)
        dados_dif.append(ld)

    df_res_b = pd.DataFrame(dados_boletim, columns=colunas_resultado)
    df_res_p = pd.DataFrame(dados_ponto, columns=colunas_resultado)
    df_res_d = pd.DataFrame(dados_dif, columns=colunas_resultado)

    registro = (
        df_boletim_filt.get("Registro", pd.Series(["-"])).iloc[0]
        if not df_boletim_filt.empty
        else "-"
    )

    boletins_set = set()
    if not df_boletim_filt.empty and "BOLETIM" in df_boletim_filt.columns:
        nums = pd.to_numeric(df_boletim_filt["BOLETIM"], errors="coerce").dropna().astype(int)
        boletins_set = set(nums.unique().tolist())

    return df_res_b, df_res_p, df_res_d, registro, boletins_set

# ============================================================
# Exportação comparativa (layout da tela + Diferenças apenas)
# ============================================================

def exportar_comparacao_individual(app):
    """Exporta comparação individual com formatação profissional"""
    pasta = filedialog.askdirectory(title="Escolha a pasta para salvar os arquivos")
    if not pasta:
        return

    try:
        need = False
        if hasattr(app.sheet_comp_boletim, "get_total_rows"):
            need = app.sheet_comp_boletim.get_total_rows() == 0
        else:
            need = not app.sheet_comp_boletim.get_sheet_data()
        if need:
            app.gerar_relatorio_comparacao()
    except Exception:
        pass

    nome_display = (app.combo_funcionario.get() or "Funcionario").strip()
    nome_file = nome_display.replace(" ", "_")
    periodo = f"{app.cal_data_inicial.entry.get()} a {app.cal_data_final.entry.get()}"

    headers_vis = _get_headers_from_sheet(app.sheet_comp_boletim) or [
        "Data", "Horas\nNormais", "Horas\nNoturnas", "Extra\n50%D", 
        "Extra\n100%D", "Extra\n50%N", "Extra\n100%N",
    ]
    headers_horas = headers_vis[1:]

    df_bol = _df_from_sheet_safe(app.sheet_comp_boletim, ["Data"] + headers_horas)
    df_pto_raw = _df_from_sheet_safe(app.sheet_comp_ponto, headers_horas)
    df_dif_raw = _df_from_sheet_safe(app.sheet_comp_diferenca, headers_horas)

    min_len = min(len(df_bol), len(df_pto_raw), len(df_dif_raw)) if len(df_bol) else 0
    df_bol = df_bol.iloc[:min_len].reset_index(drop=True)
    df_pto = pd.concat([df_bol["Data"], df_pto_raw.iloc[:min_len].reset_index(drop=True)], axis=1)
    df_dif = pd.concat([df_bol["Data"], df_dif_raw.iloc[:min_len].reset_index(drop=True)], axis=1)

    # Remove possíveis linhas "TOTAIS:"
    for df in (df_bol, df_pto, df_dif):
        if not df.empty and isinstance(df.iloc[-1, 0], str) and "TOTAIS" in df.iloc[-1, 0].upper():
            df.drop(df.index[-1], inplace=True)

    # ---- XLSX layout tela com formatação profissional ----
    caminho_xlsx = f"{pasta}/{nome_file}_comparacao.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparação"

    # Monta DataFrame final
    df_final = pd.DataFrame()
    df_final["Data"] = df_bol["Data"]
    for origem, df_src in (("Boletim", df_bol), ("Ponto", df_pto), ("Diferença", df_dif)):
        for h in headers_horas:
            df_final[f"{origem} {h.replace(chr(10), ' ')}"] = df_src[h]

    # Cabeçalho de grupos (linha 5)
    ws.append([""] + ["Boletim"] * len(headers_horas) + ["Ponto"] * len(headers_horas) + ["Diferença"] * len(headers_horas))
    
    # Cabeçalho de colunas (linha 6)
    ws.append(["Data"] + headers_horas * 3)

    # Dados
    for _, row in df_final.iterrows():
        ws.append(row.tolist())

    # Linha de totais
    totais = ["TOTAIS:"] + [somar_preservando_formato(df_final[col]) for col in df_final.columns[1:]]
    ws.append(totais)

    # Aplica cabeçalho profissional
    try:
        dt_ini = datetime.strptime(app.cal_data_inicial.entry.get(), "%d/%m/%Y")
        dt_fim = datetime.strptime(app.cal_data_final.entry.get(), "%d/%m/%Y")
    except:
        dt_ini = datetime.now()
        dt_fim = datetime.now()

    # Remove as 2 primeiras linhas para adicionar cabeçalho profissional
    ws.delete_rows(1, 2)
    
    # Insere cabeçalho profissional no início
    ws.insert_rows(1, 6)
    
    # Aplica cabeçalho
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(df_final.columns))
    titulo_cell = ws["A1"]
    titulo_cell.value = f"Relatório Comparativo – Funcionário: {nome_display}"
    titulo_cell.font = Font(name='Calibri', size=16, bold=True, color=CORES['branco'])
    titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
    titulo_cell.fill = PatternFill(start_color=CORES['primaria'], end_color=CORES['primaria'], fill_type="solid")
    ws.row_dimensions[1].height = 30
    
    # Período
    ws.row_dimensions[2].height = 22
    ws.cell(2, 1).value = f"Período: {periodo}"
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(df_final.columns))
    ws.cell(2, 1).font = Font(name='Calibri', size=11)
    ws.cell(2, 1).alignment = Alignment(horizontal="center", vertical="center")
    
    # Linhas vazias
    ws.append([])
    ws.append([])

    # Agora formata a partir da linha 7
    startrow = 7
    
    # Formata linha de grupos (linha 7)
    ws.row_dimensions[startrow].height = 25
    for c in range(1, len(df_final.columns) + 1):
        cell = ws.cell(row=startrow, column=c)
        cell.font = Font(name='Calibri', size=11, bold=True, color=CORES['branco'])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=CORES['primaria'], end_color=CORES['primaria'], fill_type="solid")
        cell.border = criar_borda('medium', CORES['primaria'])
    
    # Mescla células dos grupos
    ws.merge_cells(start_row=startrow, start_column=2, end_row=startrow, end_column=1+len(headers_horas))
    ws.merge_cells(start_row=startrow, start_column=2+len(headers_horas), end_row=startrow, end_column=1+2*len(headers_horas))
    ws.merge_cells(start_row=startrow, start_column=2+2*len(headers_horas), end_row=startrow, end_column=1+3*len(headers_horas))
    
    # Formata linha de cabeçalhos (linha 8)
    startrow += 1
    ws.row_dimensions[startrow].height = 35
    for c in range(1, len(df_final.columns) + 1):
        cell = ws.cell(row=startrow, column=c)
        cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['branco'])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color=CORES['secundaria'], end_color=CORES['secundaria'], fill_type="solid")
        cell.border = criar_borda('medium', CORES['primaria'])
    
    # Ajusta larguras
    ws.column_dimensions["A"].width = 14
    for c in range(2, len(df_final.columns) + 1):
        ws.column_dimensions[_gcl(c)].width = 12
    
    # Formata linhas de dados
    max_row = ws.max_row
    b_ini, b_fim = 2, 1 + len(headers_horas)
    p_ini, p_fim = b_fim + 1, b_fim + len(headers_horas)
    d_ini, d_fim = p_fim + 1, p_fim + len(headers_horas)
    
    for r in range(startrow + 1, max_row + 1):
        ws.row_dimensions[r].height = 20
        primeira_celula = str(ws.cell(row=r, column=1).value or "")
        is_total = _eh_total(primeira_celula)
        
        for c in range(1, len(df_final.columns) + 1):
            cell = ws.cell(row=r, column=c)
            
            if is_total:
                cell.font = Font(name='Calibri', size=11, bold=True, color=CORES['primaria'])
                cell.fill = PatternFill(start_color=CORES['neutro_medio'], end_color=CORES['neutro_medio'], fill_type="solid")
                cell.border = criar_borda('medium', CORES['primaria'])
            else:
                # Cores por seção
                if b_ini <= c <= b_fim:
                    cor_fundo = CORES['boletim']
                elif p_ini <= c <= p_fim:
                    cor_fundo = CORES['ponto']
                elif d_ini <= c <= d_fim:
                    cor_fundo = CORES['diferenca']
                else:
                    cor_fundo = CORES['branco']
                
                cell.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
                cell.border = criar_borda('thin', 'CCCCCC')
                cell.font = Font(name='Calibri', size=10)
            
            # Alinhamento
            if c == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                
                # Destaca negativos
                if eh_negativo_str(cell.value):
                    cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['erro'])
    
    ws.freeze_panes = "A9"
    wb.save(caminho_xlsx)

    # ---- PDF layout tela ----
    caminho_pdf = f"{pasta}/{nome_file}_comparacao.pdf"
    doc = SimpleDocTemplate(caminho_pdf, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=12)
    styles = getSampleStyleSheet()

    header_cells = ["Data"] + headers_horas * 3
    dados = [header_cells] + df_final.values.tolist() + [totais]

    col_widths = [1.8 * cm] + [1.36 * cm] * (len(header_cells) - 1)
    tabela = Table(dados, colWidths=col_widths, repeatRows=1)

    estilo = TableStyle([
        ("GRID", (0, 0), (-1, -1), 0.3, colors.grey),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (0, 0), (-1, 0), "CENTER"),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#183C5F")),
        ("ALIGN", (1, 1), (-1, -1), "RIGHT"),
        ("FONTSIZE", (0, 0), (-1, -1), 7),
    ])
    
    estilo.add("BACKGROUND", (b_ini, 1), (b_fim, -1), colors.HexColor("#B7DFFB"))
    estilo.add("BACKGROUND", (p_ini, 1), (p_fim, -1), colors.HexColor("#C5EDC1"))
    estilo.add("BACKGROUND", (d_ini, 1), (d_fim, -1), colors.HexColor("#FFDAB9"))

    # Negativos em diferença
    for r in range(1, len(dados)):
        for c in range(d_ini, d_fim + 1):
            if eh_negativo_str(dados[r][c]):
                estilo.add("TEXTCOLOR", (c, r), (c, r), colors.red)
                estilo.add("FONTNAME", (c, r), (c, r), "Helvetica-Bold")

    tabela.setStyle(estilo)
    story = [
        Paragraph("<b>Relatório Comparativo – Funcionário Atual</b>", styles["Title"]),
        Paragraph(f"<b>Funcionário:</b> {nome_display}", styles["Normal"]),
        Paragraph(f"<b>Período:</b> {periodo}", styles["Normal"]),
        Spacer(1, 8),
        tabela,
    ]
    doc.build(story)

    messagebox.showinfo("Sucesso", f"Arquivos salvos:\n- {caminho_xlsx}\n- {caminho_pdf}")

def exportar_comparacao_unificada(app):
    """Exporta layout da tela + cria um par XLSX/PDF apenas com as DIFERENÇAS com formatação profissional"""
    nome_display = (app.combo_funcionario.get() or "Funcionario").strip()
    arquivo_base = filedialog.asksaveasfilename(
        defaultextension="",
        filetypes=[("Todos os Arquivos", "*.*")],
        title="Salvar Exportação Unificada",
        initialfile=f"Comparacao_{nome_display.replace(' ', '_')}",
    )
    if not arquivo_base:
        return

    caminho_xlsx = arquivo_base + ".xlsx"
    caminho_pdf = arquivo_base + ".pdf"

    headers_vis = _get_headers_from_sheet(app.sheet_comp_boletim) or [
        "Data", "Horas\nNormais", "Horas\nNoturnas", "Extra\n50%D",
        "Extra\n100%D", "Extra\n50%N", "Extra\n100%N",
    ]
    headers = headers_vis[1:]

    df_bol = _df_from_sheet_safe(app.sheet_comp_boletim, ["Data"] + headers)
    df_pto_raw = _df_from_sheet_safe(app.sheet_comp_ponto, headers)
    df_dif_raw = _df_from_sheet_safe(app.sheet_comp_diferenca, headers)

    min_len = min(len(df_bol), len(df_pto_raw), len(df_dif_raw)) if len(df_bol) else 0
    df_bol = df_bol.iloc[:min_len].reset_index(drop=True)
    df_pto = pd.concat([df_bol["Data"], df_pto_raw.iloc[:min_len].reset_index(drop=True)], axis=1)
    df_dif = pd.concat([df_bol["Data"], df_dif_raw.iloc[:min_len].reset_index(drop=True)], axis=1)

    for df in (df_bol, df_pto, df_dif):
        if not df.empty and isinstance(df.iloc[-1, 0], str) and "TOTAIS" in df.iloc[-1, 0].upper():
            df.drop(df.index[-1], inplace=True)

    # Monta DF final
    df_final = pd.DataFrame()
    df_final["Data"] = df_bol["Data"]
    for origem, df_src in (("Boletim", df_bol), ("Ponto", df_pto), ("Diferença", df_dif)):
        for h in headers:
            df_final[f"{origem} {h.replace(chr(10), ' ')}"] = df_src[h]

    totais = ["TOTAIS:"] + [somar_preservando_formato(df_final[col]) for col in df_final.columns[1:]]

    # ========= XLSX layout tela com formatação profissional =========
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparação"
    
    periodo = f"{app.cal_data_inicial.entry.get()} a {app.cal_data_final.entry.get()}"
    
    try:
        dt_ini = datetime.strptime(app.cal_data_inicial.entry.get(), "%d/%m/%Y")
        dt_fim = datetime.strptime(app.cal_data_final.entry.get(), "%d/%m/%Y")
    except:
        dt_ini = datetime.now()
        dt_fim = datetime.now()
    
    # Aplica cabeçalho profissional
    aplicar_cabecalho_profissional(
        ws,
        titulo=f"Relatório Comparativo Unificado – Funcionário: {nome_display}",
        contrato="-",
        dt_ini=dt_ini,
        dt_fim=dt_fim,
        boletins="-",
        df_cols=list(df_final.columns)
    )
    
    # Linha de grupos (linha 7)
    startrow = STARTROW_DADOS
    ws.cell(row=startrow, column=1).value = ""
    for i, grupo in enumerate(["Boletim", "Ponto", "Diferença"], start=1):
        ws.cell(row=startrow, column=1 + (i-1)*len(headers) + 1).value = grupo
    
    # Mescla células dos grupos
    ws.merge_cells(start_row=startrow, start_column=2, end_row=startrow, end_column=1+len(headers))
    ws.merge_cells(start_row=startrow, start_column=2+len(headers), end_row=startrow, end_column=1+2*len(headers))
    ws.merge_cells(start_row=startrow, start_column=2+2*len(headers), end_row=startrow, end_column=1+3*len(headers))
    
    # Formata linha de grupos
    ws.row_dimensions[startrow].height = 25
    for c in range(1, len(df_final.columns) + 1):
        cell = ws.cell(row=startrow, column=c)
        cell.font = Font(name='Calibri', size=11, bold=True, color=CORES['branco'])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=CORES['primaria'], end_color=CORES['primaria'], fill_type="solid")
        cell.border = criar_borda('medium', CORES['primaria'])
    
    # Cabeçalhos de colunas (linha 8)
    startrow += 1
    ws.cell(row=startrow, column=1).value = "Data"
    col_idx = 2
    for _ in range(3):  # Boletim, Ponto, Diferença
        for h in headers:
            ws.cell(row=startrow, column=col_idx).value = h
            col_idx += 1
    
    # Formata cabeçalhos
    ws.row_dimensions[startrow].height = 35
    for c in range(1, len(df_final.columns) + 1):
        cell = ws.cell(row=startrow, column=c)
        cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['branco'])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color=CORES['secundaria'], end_color=CORES['secundaria'], fill_type="solid")
        cell.border = criar_borda('medium', CORES['primaria'])
    
    # Dados
    for _, row in df_final.iterrows():
        startrow += 1
        for col_idx, val in enumerate(row, start=1):
            ws.cell(row=startrow, column=col_idx).value = val
    
    # Totais
    startrow += 1
    for col_idx, val in enumerate(totais, start=1):
        ws.cell(row=startrow, column=col_idx).value = val
    
    # Formata dados e totais
    b_ini, b_fim = 2, 1 + len(headers)
    p_ini, p_fim = b_fim + 1, b_fim + len(headers)
    d_ini, d_fim = p_fim + 1, p_fim + len(headers)
    
    max_row = ws.max_row
    for r in range(STARTROW_DADOS + 2, max_row + 1):
        ws.row_dimensions[r].height = 20
        primeira_celula = str(ws.cell(row=r, column=1).value or "")
        is_total = _eh_total(primeira_celula)
        
        for c in range(1, len(df_final.columns) + 1):
            cell = ws.cell(row=r, column=c)
            
            if is_total:
                cell.font = Font(name='Calibri', size=11, bold=True, color=CORES['primaria'])
                cell.fill = PatternFill(start_color=CORES['neutro_medio'], end_color=CORES['neutro_medio'], fill_type="solid")
                cell.border = criar_borda('medium', CORES['primaria'])
            else:
                if b_ini <= c <= b_fim:
                    cor_fundo = CORES['boletim']
                elif p_ini <= c <= p_fim:
                    cor_fundo = CORES['ponto']
                elif d_ini <= c <= d_fim:
                    cor_fundo = CORES['diferenca']
                else:
                    cor_fundo = CORES['branco']
                
                cell.fill = PatternFill(start_color=cor_fundo, end_color=cor_fundo, fill_type="solid")
                cell.border = criar_borda('thin', 'CCCCCC')
                cell.font = Font(name='Calibri', size=10)
            
            if c == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if eh_negativo_str(cell.value):
                    cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['erro'])
    
    # Ajusta larguras
    ws.column_dimensions["A"].width = 14
    for c in range(2, len(df_final.columns) + 1):
        ws.column_dimensions[_gcl(c)].width = 12
    
    ws.freeze_panes = f"A{STARTROW_DADOS + 2}"
    wb.save(caminho_xlsx)

    # ========= DIFERENÇAS APENAS com formatação profissional =========
    caminho_xlsx_dif = arquivo_base + "_Diferencas.xlsx"
    caminho_pdf_dif = arquivo_base + "_Diferencas.pdf"

    totais_dif = ["TOTAIS:"] + [somar_preservando_formato(df_dif[h]) for h in headers]

    wb_d = Workbook()
    ws_d = wb_d.active
    ws_d.title = "Diferenças"
    
    aplicar_cabecalho_profissional(
        ws_d,
        titulo=f"Relatório – Diferenças (Boletim − Ponto) – Funcionário: {nome_display}",
        contrato="-",
        dt_ini=dt_ini,
        dt_fim=dt_fim,
        boletins="-",
        df_cols=["Data"] + headers
    )
    
    # Linha de grupo (linha 7)
    ws_d.cell(row=STARTROW_DADOS, column=1).value = ""
    ws_d.merge_cells(start_row=STARTROW_DADOS, start_column=2, end_row=STARTROW_DADOS, end_column=1+len(headers))
    ws_d.cell(row=STARTROW_DADOS, column=2).value = "Diferença"
    
    ws_d.row_dimensions[STARTROW_DADOS].height = 25
    for c in range(1, 2 + len(headers)):
        cell = ws_d.cell(row=STARTROW_DADOS, column=c)
        cell.font = Font(name='Calibri', size=11, bold=True, color=CORES['branco'])
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill(start_color=CORES['primaria'], end_color=CORES['primaria'], fill_type="solid")
        cell.border = criar_borda('medium', CORES['primaria'])
    
    # Cabeçalhos (linha 8)
    startrow_d = STARTROW_DADOS + 1
    ws_d.cell(row=startrow_d, column=1).value = "Data"
    for col_idx, h in enumerate(headers, start=2):
        ws_d.cell(row=startrow_d, column=col_idx).value = h
    
    ws_d.row_dimensions[startrow_d].height = 35
    for c in range(1, 2 + len(headers)):
        cell = ws_d.cell(row=startrow_d, column=c)
        cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['branco'])
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.fill = PatternFill(start_color=CORES['secundaria'], end_color=CORES['secundaria'], fill_type="solid")
        cell.border = criar_borda('medium', CORES['primaria'])
    
    # Dados
    for _, row in df_dif.iterrows():
        startrow_d += 1
        for col_idx, val in enumerate(row, start=1):
            ws_d.cell(row=startrow_d, column=col_idx).value = val
    
    # Totais
    startrow_d += 1
    for col_idx, val in enumerate(totais_dif, start=1):
        ws_d.cell(row=startrow_d, column=col_idx).value = val
    
    # Formata dados
    max_row_d = ws_d.max_row
    for r in range(STARTROW_DADOS + 2, max_row_d + 1):
        ws_d.row_dimensions[r].height = 20
        primeira_celula = str(ws_d.cell(row=r, column=1).value or "")
        is_total = _eh_total(primeira_celula)
        
        for c in range(1, 2 + len(headers)):
            cell = ws_d.cell(row=r, column=c)
            
            if is_total:
                cell.font = Font(name='Calibri', size=11, bold=True, color=CORES['primaria'])
                cell.fill = PatternFill(start_color=CORES['neutro_medio'], end_color=CORES['neutro_medio'], fill_type="solid")
                cell.border = criar_borda('medium', CORES['primaria'])
            else:
                cell.fill = PatternFill(start_color=CORES['diferenca'], end_color=CORES['diferenca'], fill_type="solid")
                cell.border = criar_borda('thin', 'CCCCCC')
                cell.font = Font(name='Calibri', size=10)
            
            if c == 1:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="right", vertical="center")
                if eh_negativo_str(cell.value):
                    cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['erro'])
    
    ws_d.column_dimensions["A"].width = 14
    for c in range(2, 2 + len(headers)):
        ws_d.column_dimensions[_gcl(c)].width = 12
    
    ws_d.freeze_panes = f"A{STARTROW_DADOS + 2}"
    wb_d.save(caminho_xlsx_dif)

    # PDF diferenças
    doc_d = SimpleDocTemplate(caminho_pdf_dif, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=12)
    styles = getSampleStyleSheet()
    header_cells_d = ["Data"] + headers
    dados_d = [header_cells_d] + df_dif.values.tolist() + [totais_dif]
    tabela_d = Table(dados_d, colWidths=[1.8*cm] + [1.36*cm]*(len(header_cells_d)-1), repeatRows=1)
    estilo_d = TableStyle([
        ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#183C5F")),
        ("ALIGN", (1,1), (-1,-1), "RIGHT"),
        ("FONTSIZE", (0,0), (-1,-1), 7),
    ])
    estilo_d.add("BACKGROUND", (1,1), (-1,-1), colors.HexColor("#FFDAB9"))
    for r in range(1, len(dados_d)):
        for c in range(1, len(header_cells_d)):
            if eh_negativo_str(dados_d[r][c]):
                estilo_d.add("TEXTCOLOR", (c, r), (c, r), colors.red)
                estilo_d.add("FONTNAME", (c, r), (c, r), "Helvetica-Bold")
    tabela_d.setStyle(estilo_d)

    story_d = [
        Paragraph("<b>Relatório – Diferenças (Boletim − Ponto)</b>", styles["Title"]),
        Paragraph(f"<b>Funcionário:</b> {nome_display}", styles["Normal"]),
        Paragraph(f"<b>Período:</b> {periodo}", styles["Normal"]),
        Spacer(1, 8),
        tabela_d,
    ]
    doc_d.build(story_d)

    # PDF layout tela
    doc = SimpleDocTemplate(caminho_pdf, pagesize=landscape(A4), rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=12)
    styles = getSampleStyleSheet()
    header_cells = ["Data"] + headers * 3
    dados = [header_cells] + df_final.values.tolist() + [totais]
    tabela = Table(dados, colWidths=[1.8*cm] + [1.36*cm]*(len(header_cells)-1), repeatRows=1)
    estilo = TableStyle([
        ("GRID", (0,0), (-1,-1), 0.3, colors.grey),
        ("VALIGN", (0,0), (-1,-1), "MIDDLE"),
        ("ALIGN", (0,0), (-1,0), "CENTER"),
        ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
        ("TEXTCOLOR", (0,0), (-1,0), colors.whitesmoke),
        ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#183C5F")),
        ("ALIGN", (1,1), (-1,-1), "RIGHT"),
        ("FONTSIZE", (0,0), (-1,-1), 7),
    ])
    
    estilo.add("BACKGROUND", (b_ini,1), (b_fim,-1), colors.HexColor("#B7DFFB"))
    estilo.add("BACKGROUND", (p_ini,1), (p_fim,-1), colors.HexColor("#C5EDC1"))
    estilo.add("BACKGROUND", (d_ini,1), (d_fim,-1), colors.HexColor("#FFDAB9"))
    for r in range(1, len(dados)):
        for c in range(d_ini, d_fim + 1):
            if eh_negativo_str(dados[r][c]):
                estilo.add("TEXTCOLOR", (c,r), (c,r), colors.red)
                estilo.add("FONTNAME", (c,r), (c,r), "Helvetica-Bold")
    tabela.setStyle(estilo)

    story = [
        Paragraph("<b>Relatório Comparativo Unificado</b>", styles["Title"]),
        Paragraph(f"<b>Funcionário:</b> {nome_display}", styles["Normal"]),
        Paragraph(f"<b>Período:</b> {periodo}", styles["Normal"]),
        Spacer(1, 8),
        tabela,
    ]
    doc.build(story)

    messagebox.showinfo(
        "Sucesso",
        "Exportações concluídas!\n\n"
        f"XLSX: {caminho_xlsx}\nPDF:  {caminho_pdf}\n\n"
        f"XLSX (Diferenças): {caminho_xlsx_dif}\nPDF  (Diferenças): {caminho_pdf_dif}"
    )

# ============================================================
# Comparação geral e Totais Mensais
# ============================================================

def exportar_comparacao_geral(app):
    try:
        pasta = filedialog.askdirectory(title="Escolha uma pasta para salvar os arquivos")
        if not pasta:
            return
        if app.dados_df is None or app.df_ponto is None:
            messagebox.showwarning("Aviso", "Dados completos não carregados.")
            return
        hoje = dt.datetime.now().strftime("%Y-%m-%d")
        app.dados_df.to_excel(f"{pasta}/boletim_completo_{hoje}.xlsx", index=False)
        app.df_ponto.to_excel(f"{pasta}/ponto_completo_{hoje}.xlsx", index=False)
        doc = SimpleDocTemplate(f"{pasta}/comparacao_geral_{hoje}.pdf", pagesize=landscape(A4))
        styles = getSampleStyleSheet()
        story = [
            Paragraph("Comparação Geral de Dados", styles["Title"]),
            Paragraph(f"Exportado em: {hoje}", styles["Normal"]),
            Spacer(1, 0.3 * inch),
            Paragraph("⚠️ PDF não contém tabelas completas. Verifique os .xlsx gerados.", styles["Normal"]),
        ]
        doc.build(story)
        messagebox.showinfo("Sucesso", f"Dados completos exportados para: {pasta}")
    except Exception as e:
        messagebox.showerror("Erro", str(e))

def exportar_totais_mensais(df_group: pd.DataFrame, nome_base: str, pasta_destino: str):
    try:
        caminho_xlsx = f"{pasta_destino}/{nome_base}.xlsx"
        caminho_pdf = f"{pasta_destino}/{nome_base}.pdf"
        
        # Excel com formatação profissional
        wb = Workbook()
        ws = wb.active
        ws.title = "Totais Mensais"
        
        # Escreve cabeçalhos
        for col_idx, col_name in enumerate(df_group.columns, start=1):
            ws.cell(row=STARTROW_DADOS, column=col_idx, value=col_name)
        
        # Escreve dados
        for row_idx, (_, row) in enumerate(df_group.iterrows(), start=STARTROW_DADOS + 1):
            for col_idx, col_name in enumerate(df_group.columns, start=1):
                ws.cell(row=row_idx, column=col_idx, value=row[col_name])
        
        # Aplica cabeçalho profissional
        try:
            ano_mes = nome_base[-6:]
            ano = int(ano_mes[:4])
            mes = int(ano_mes[4:])
            dt_ref = datetime(ano, mes, 1)
        except:
            dt_ref = datetime.now()
        
        aplicar_cabecalho_profissional(
            ws,
            titulo=f"Totais Mensais por Funcionário - {ano_mes}",
            contrato="-",
            dt_ini=dt_ref,
            dt_fim=dt_ref,
            boletins="-",
            df_cols=list(df_group.columns)
        )
        
        formatar_planilha_profissional(ws, startrow=STARTROW_DADOS)
        wb.save(caminho_xlsx)

        # PDF
        dados = [list(df_group.columns)] + df_group.values.tolist()
        for i in range(1, len(dados)):
            for j in range(3, len(dados[i])):
                try:
                    dados[i][j] = f"{float(dados[i][j]):.2f}"
                except Exception:
                    pass

        doc = SimpleDocTemplate(caminho_pdf, pagesize=A4)
        styles = getSampleStyleSheet()
        tabela = Table(dados, repeatRows=1)
        tabela.setStyle(
            TableStyle([
                ("BACKGROUND", (0,0), (-1,0), colors.HexColor("#003366")),
                ("TEXTCOLOR", (0,0), (-1,0), colors.white),
                ("ALIGN", (0,0), (-1,-1), "CENTER"),
                ("FONTNAME", (0,0), (-1,0), "Helvetica-Bold"),
                ("FONTSIZE", (0,0), (-1,0), 6),
                ("FONTSIZE", (0,1), (-1,-1), 6),
                ("BACKGROUND", (0,1), (-1,-1), colors.whitesmoke),
                ("GRID", (0,0), (-1,-1), 0.5, colors.grey),
            ])
        )
        doc.build([
            Paragraph(f"Totais Mensais por Funcionário - {ano_mes}", styles["Title"]),
            Spacer(1, 12),
            tabela,
        ])
        messagebox.showinfo("Exportado", f"Totais mensais salvos:\n- {nome_base}.xlsx\n- {nome_base}.pdf")
    except Exception as e:
        messagebox.showerror("Erro", f"Falha na exportação de totais mensais:\n{e}")

# ============================================================
# Consolidação por Contrato (com formatação profissional)
# ============================================================

def exportar_totais_consolidados_excel(app, lista_contratos):
    """
    Cria um workbook com formatação profissional:
      • Abas por contrato: funcionários pertencentes ao contrato e seus totais
      • Aba "Resumo Geral": Nome | Contrato | Boletim | totais…
      • Um segundo workbook "_Diferencas.xlsx" apenas com diferenças
    """
    arquivo_saida = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Arquivos Excel", "*.xlsx")],
        title="Salvar Totais Consolidados por Contrato",
        initialfile="Totais_Consolidados_por_Contrato.xlsx",
    )
    if not arquivo_saida:
        return

    # Período da UI
    dt_ini = datetime.strptime(app.cal_data_inicial.entry.get(), "%d/%m/%Y")
    dt_fim = datetime.strptime(app.cal_data_final.entry.get(), "%d/%m/%Y")

    # Calcula totais por funcionário
    df_boletim_filtrado = app.dados_df[app.dados_df["Contrato"].isin(lista_contratos)]
    todos_funcionarios = sorted(df_boletim_filtrado["Funcionário"].dropna().unique())

    resultados = []
    for nome in todos_funcionarios:
        totais = app._calcular_totais_funcionario(nome)
        if not totais:
            continue
        contratos_do_func = df_boletim_filtrado[df_boletim_filtrado["Funcionário"] == nome]["Contrato"].unique()
        totais["Contratos"] = ", ".join(map(str, contratos_do_func))
        resultados.append(totais)

    if not resultados:
        messagebox.showwarning("Aviso", "Nenhum dado encontrado para os critérios selecionados.")
        return

    df_all = pd.DataFrame(resultados)

    with pd.ExcelWriter(arquivo_saida, engine="openpyxl") as writer:
        used_names = set()
        
        # Cria aba por contrato
        for contrato_id in sorted(set(lista_contratos)):
            sheet_name = _sheet_name_unique(contrato_id, used_names)
            df_contrato = df_all[df_all["Contratos"].apply(lambda x: str(contrato_id) in str(x))].copy()
            if df_contrato.empty:
                continue

            df_contrato.drop(columns=["Contratos"], inplace=True, errors="ignore")
            if "Funcionário" in df_contrato.columns:
                df_contrato = df_contrato.sort_values("Funcionário", kind="stable").reset_index(drop=True)

            # Linha de totais
            num_cols = list(df_contrato.select_dtypes(include="number").columns)
            linha_total = {c: "" for c in df_contrato.columns}
            for c in num_cols:
                try:
                    linha_total[c] = round(float(pd.to_numeric(df_contrato[c], errors="coerce").fillna(0).sum()), 2)
                except Exception:
                    linha_total[c] = pd.to_numeric(df_contrato[c], errors="coerce").fillna(0).sum()
            
            for c in df_contrato.columns:
                if c not in num_cols and c not in ("Funcionário", "Registro", "Contrato", "Boletim"):
                    try:
                        linha_total[c] = somar_preservando_formato(df_contrato[c])
                    except Exception:
                        pass
            
            if "Funcionário" in linha_total:
                linha_total["Funcionário"] = "TOTAIS"
            
            df_contrato = pd.concat([df_contrato, pd.DataFrame([linha_total])], ignore_index=True)

            # Arredonda numéricas
            if num_cols:
                df_contrato[num_cols] = df_contrato[num_cols].apply(lambda s: s.round(2))

            df_contrato.columns = [str(c).replace("_", " ") for c in df_contrato.columns]
            
            # Escreve na planilha
            df_contrato.to_excel(writer, sheet_name=sheet_name, index=False, startrow=STARTROW_DADOS)

            # Obtém worksheet e aplica formatação profissional
            ws = writer.sheets[sheet_name]
            
            # Busca boletins do contrato
            df_b = app.dados_df.copy()
            df_b["DATA"] = pd.to_datetime(df_b["DATA"], errors="coerce")
            mask = (
                (df_b["Contrato"] == contrato_id)
                & (df_b["DATA"].dt.date >= dt_ini.date())
                & (df_b["DATA"].dt.date <= dt_fim.date())
            )
            df_b = df_b.loc[mask]
            if "BOLETIM" in df_b.columns:
                nums = (
                    pd.to_numeric(df_b["BOLETIM"], errors="coerce")
                    .dropna().astype(int).sort_values().unique().tolist()
                )
                boletins_txt = ", ".join(map(str, nums)) if nums else "-"
            else:
                boletins_txt = "-"

            # Aplica cabeçalho e formatação profissional
            aplicar_cabecalho_profissional(
                ws,
                titulo="Relatório Consolidado de Totais",
                contrato=contrato_id,
                dt_ini=dt_ini,
                dt_fim=dt_fim,
                boletins=boletins_txt,
                df_cols=list(df_contrato.columns),
            )
            formatar_planilha_profissional(ws, startrow=STARTROW_DADOS)

        # ---- Resumo Geral ----
        df_resumo = df_all.copy()
        df_resumo["Contrato"] = df_resumo.get("Contratos", "").apply(_ult5_contratos)

        # Boletins por funcionário
        boletins_por_func = {}
        df_b_all = app.dados_df.copy()
        df_b_all["DATA"] = pd.to_datetime(df_b_all["DATA"], errors="coerce")
        mask_all = (
            df_b_all["DATA"].dt.date.between(dt_ini.date(), dt_fim.date())
            & df_b_all["Contrato"].isin(lista_contratos)
        )
        df_b_all = df_b_all.loc[mask_all]
        if "BOLETIM" in df_b_all.columns:
            for nome in df_resumo["Funcionário"].dropna().unique():
                nums = (
                    pd.to_numeric(df_b_all.loc[df_b_all["Funcionário"] == nome, "BOLETIM"], errors="coerce")
                    .dropna().astype(int).sort_values().unique().tolist()
                )
                boletins_por_func[nome] = ", ".join(map(str, nums)) if nums else ""
        df_resumo["Boletim"] = df_resumo["Funcionário"].map(boletins_por_func).fillna("")

        cols_rest = [c for c in df_resumo.columns if c not in ("Funcionário", "Contratos", "Contrato", "Boletim")]
        df_resumo = df_resumo[["Funcionário", "Contrato", "Boletim"] + cols_rest]
        if "Funcionário" in df_resumo.columns:
            df_resumo = df_resumo.sort_values("Funcionário", kind="stable").reset_index(drop=True)

        # Totais mistos
        num_cols = list(df_resumo.select_dtypes(include="number").columns)
        linha_total = {c: "" for c in df_resumo.columns}
        if num_cols:
            for c in num_cols:
                try:
                    linha_total[c] = round(float(pd.to_numeric(df_resumo[c], errors="coerce").fillna(0).sum()), 2)
                except Exception:
                    linha_total[c] = pd.to_numeric(df_resumo[c], errors="coerce").fillna(0).sum()
        if "Funcionário" in linha_total:
            linha_total["Funcionário"] = "TOTAIS GERAIS"
        linha_total["Contrato"] = ""
        linha_total["Boletim"] = ""
        df_resumo = pd.concat([df_resumo, pd.DataFrame([linha_total])], ignore_index=True)

        df_resumo.columns = [str(c).replace("_", " ") for c in df_resumo.columns]
        df_resumo.to_excel(writer, sheet_name="Resumo Geral", index=False, startrow=STARTROW_DADOS)

        ws_resumo = writer.sheets["Resumo Geral"]
        
        # Boletins gerais
        if "BOLETIM" in df_b_all.columns:
            all_nums = pd.to_numeric(df_b_all["BOLETIM"], errors="coerce").dropna().astype(int).sort_values().unique().tolist()
            boletins_all_text = ", ".join(map(str, all_nums)) if all_nums else "-"
        else:
            boletins_all_text = "-"

        aplicar_cabecalho_profissional(
            ws_resumo,
            titulo="Resumo Geral - Totais por Funcionário",
            contrato=", ".join(map(str, lista_contratos)),
            dt_ini=dt_ini,
            dt_fim=dt_fim,
            boletins=boletins_all_text,
            df_cols=list(df_resumo.columns),
        )
        formatar_planilha_profissional(ws_resumo, startrow=STARTROW_DADOS)

    # ================= Workbook apenas de DIFERENÇAS =================
    diff_cols_raw = [c for c in df_all.columns if str(c).startswith("Diferença ")]
    if diff_cols_raw:
        arquivo_dif = (arquivo_saida[:-5] + "_Diferencas.xlsx") if arquivo_saida.lower().endswith(".xlsx") else (arquivo_saida + "_Diferencas.xlsx")
        with pd.ExcelWriter(arquivo_dif, engine="openpyxl") as writer:
            used_d = set()
            for contrato_id in sorted(set(lista_contratos)):
                sheet_name = _sheet_name_unique(contrato_id, used_d)
                df_contrato = df_all[df_all["Contratos"].apply(lambda x: str(contrato_id) in str(x))].copy()
                if df_contrato.empty:
                    continue
                df_dif = df_contrato[["Funcionário"] + diff_cols_raw].copy()
                df_dif.columns = [str(c).replace("_", " ") for c in df_dif.columns]
                somas = df_dif.select_dtypes(include="number").sum()
                linha_total = pd.Series(somas, name="TOTAIS")
                linha_total["Funcionário"] = "TOTAIS"
                df_dif = pd.concat([df_dif, linha_total.to_frame().T], ignore_index=True)

                df_dif.to_excel(writer, sheet_name=sheet_name, index=False, startrow=STARTROW_DADOS)
                ws = writer.sheets[sheet_name]

                # Aplica cabeçalho e formatação profissional
                aplicar_cabecalho_profissional(
                    ws,
                    titulo="Relatório – DIFERENÇAS (Boletim − Ponto)",
                    contrato=contrato_id,
                    dt_ini=dt_ini,
                    dt_fim=dt_fim,
                    boletins="-",
                    df_cols=list(df_dif.columns),
                )
                formatar_planilha_profissional(ws, startrow=STARTROW_DADOS)

                # Destaca negativos em vermelho
                max_row, max_col = ws.max_row, ws.max_column
                for r in range(STARTROW_DADOS + 1, max_row + 1):
                    for c in range(2, max_col + 1):
                        cell = ws.cell(row=r, column=c)
                        if eh_negativo_str(cell.value):
                            cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['erro'])

            # Resumo Geral (diferenças)
            df_res = df_all.copy()
            df_res["Contrato"] = df_res.get("Contratos", "").apply(_ult5_contratos)
            keep = ["Funcionário", "Contrato"] + diff_cols_raw
            df_res = df_res[keep]
            df_res.columns = [str(c).replace("_", " ") for c in df_res.columns]

            num_cols = df_res.select_dtypes(include="number").columns
            tot = df_res[num_cols].sum(numeric_only=True)
            linha_total = pd.Series(index=df_res.columns, dtype=object)
            linha_total["Funcionário"] = "TOTAIS GERAIS"
            linha_total["Contrato"] = ""
            for c in num_cols:
                try:
                    linha_total[c] = round(float(tot.get(c, 0) or 0), 2)
                except Exception:
                    linha_total[c] = tot.get(c, 0)
            df_res = pd.concat([df_res, linha_total.to_frame().T], ignore_index=True)

            df_res.to_excel(writer, sheet_name="Resumo Geral", index=False, startrow=STARTROW_DADOS)
            ws = writer.sheets["Resumo Geral"]
            
            aplicar_cabecalho_profissional(
                ws,
                titulo="Resumo Geral – DIFERENÇAS (Boletim − Ponto)",
                contrato=", ".join(map(str, lista_contratos)),
                dt_ini=dt_ini,
                dt_fim=dt_fim,
                boletins="-",
                df_cols=list(df_res.columns),
            )
            formatar_planilha_profissional(ws, startrow=STARTROW_DADOS)

            # Destaca negativos
            max_row, max_col = ws.max_row, ws.max_column
            for r in range(STARTROW_DADOS + 1, max_row + 1):
                for c in range(3, max_col + 1):
                    cell = ws.cell(row=r, column=c)
                    if eh_negativo_str(cell.value):
                        cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['erro'])

    messagebox.showinfo("Sucesso", f"Relatório de totais consolidados salvo em:\n{arquivo_saida}")

def exportar_totais_consolidados_por_contrato_em_arquivos(app, lista_contratos):
    """
    Gera UM PAR de arquivos por contrato com formatação profissional:
    • Totais_Consolidados_<contrato>.xlsx
    • Totais_Consolidados_<contrato>_Diferencas.xlsx
    """
    pasta = filedialog.askdirectory(title="Escolha a pasta para salvar os arquivos por contrato")
    if not pasta:
        return

    # Período da UI
    try:
        dt_ini = datetime.strptime(app.cal_data_inicial.entry.get(), "%d/%m/%Y")
        dt_fim = datetime.strptime(app.cal_data_final.entry.get(), "%d/%m/%Y")
    except Exception:
        messagebox.showerror("Erro", "Período inválido na interface.")
        return

    # Prepara base de totais por funcionário
    df_boletim_sel = app.dados_df[app.dados_df["Contrato"].isin(lista_contratos)].copy()
    funcionarios = sorted(df_boletim_sel["Funcionário"].dropna().unique())

    resultados = []
    for nome in funcionarios:
        totais = app._calcular_totais_funcionario(nome)
        if not totais:
            continue
        contratos_do_func = df_boletim_sel[df_boletim_sel["Funcionário"] == nome]["Contrato"].unique()
        totais["Contratos"] = ", ".join(map(str, contratos_do_func))
        resultados.append(totais)

    if not resultados:
        messagebox.showwarning("Aviso", "Nenhum dado encontrado para os contratos selecionados.")
        return

    df_all = pd.DataFrame(resultados)
    gerados = []

    # Loop por contrato → 2 arquivos
    for contrato_id in sorted(set(lista_contratos)):
        # ---- Totais ----
        df_contrato = df_all[df_all["Contratos"].apply(lambda x: str(contrato_id) in str(x))].copy()
        if not df_contrato.empty:
            df_contrato.drop(columns=["Contratos"], inplace=True, errors="ignore")
            
            # Linha TOTAIS
            somas = df_contrato.select_dtypes(include="number").sum()
            linha_total = pd.Series(somas, name="TOTAIS")
            linha_total["Funcionário"] = "TOTAIS"
            df_contrato = pd.concat([df_contrato, linha_total.to_frame().T], ignore_index=True)

            df_contrato.columns = [str(c).replace("_", " ") for c in df_contrato.columns]

            # Cria workbook e escreve
            wb = Workbook()
            ws = wb.active
            ws.title = str(contrato_id)

            # Escrevendo a partir de STARTROW_DADOS
            for j, col in enumerate(df_contrato.columns, start=1):
                ws.cell(row=STARTROW_DADOS, column=j, value=col)
            for i, (_, row) in enumerate(df_contrato.iterrows(), start=STARTROW_DADOS + 1):
                for j, col in enumerate(df_contrato.columns, start=1):
                    ws.cell(row=i, column=j, value=row[col])

            # Boletins do contrato no período
            df_b = app.dados_df.copy()
            df_b["DATA"] = pd.to_datetime(df_b["DATA"], errors="coerce")
            mask = (df_b["Contrato"] == contrato_id) & (df_b["DATA"].dt.date.between(dt_ini.date(), dt_fim.date()))
            df_b = df_b.loc[mask]
            if "BOLETIM" in df_b.columns:
                boletins_txt = ", ".join(
                    map(str, pd.to_numeric(df_b["BOLETIM"], errors="coerce").dropna().astype(int).sort_values().unique().tolist())
                ) or "-"
            else:
                boletins_txt = "-"

            aplicar_cabecalho_profissional(
                ws,
                titulo="Relatório Consolidado de Totais",
                contrato=contrato_id,
                dt_ini=dt_ini,
                dt_fim=dt_fim,
                boletins=boletins_txt,
                df_cols=list(df_contrato.columns),
            )
            formatar_planilha_profissional(ws, startrow=STARTROW_DADOS)

            # Salvar
            nome_tot = _safe_filename(f"Totais_Consolidados_{contrato_id}.xlsx")
            caminho_tot = f"{pasta}/{nome_tot}"
            wb.save(caminho_tot)
            gerados.append(caminho_tot)

        # ---- Diferenças ----
        diff_cols = [c for c in df_all.columns if str(c).startswith("Diferença ")]
        if diff_cols:
            df_dif = df_all[df_all["Contratos"].apply(lambda x: str(contrato_id) in str(x))].copy()
            if not df_dif.empty:
                df_dif = df_dif[["Funcionário"] + diff_cols].copy()
                df_dif.columns = [str(c).replace("_", " ") for c in df_dif.columns]
                
                # Linha TOTAIS
                somas = df_dif.select_dtypes(include="number").sum()
                linha_total = pd.Series(somas, name="TOTAIS")
                linha_total["Funcionário"] = "TOTAIS"
                df_dif = pd.concat([df_dif, linha_total.to_frame().T], ignore_index=True)

                wb_d = Workbook()
                ws_d = wb_d.active
                ws_d.title = str(contrato_id)
                
                for j, col in enumerate(df_dif.columns, start=1):
                    ws_d.cell(row=STARTROW_DADOS, column=j, value=col)
                for i, (_, row) in enumerate(df_dif.iterrows(), start=STARTROW_DADOS + 1):
                    for j, col in enumerate(df_dif.columns, start=1):
                        ws_d.cell(row=i, column=j, value=row[col])

                aplicar_cabecalho_profissional(
                    ws_d,
                    titulo="Relatório – DIFERENÇAS (Boletim − Ponto)",
                    contrato=contrato_id,
                    dt_ini=dt_ini,
                    dt_fim=dt_fim,
                    boletins="-",
                    df_cols=list(df_dif.columns),
                )
                formatar_planilha_profissional(ws_d, startrow=STARTROW_DADOS)

                # Negativos em vermelho + negrito
                max_row, max_col = ws_d.max_row, ws_d.max_column
                for r in range(STARTROW_DADOS + 1, max_row + 1):
                    for c in range(2, max_col + 1):
                        cell = ws_d.cell(row=r, column=c)
                        if eh_negativo_str(cell.value):
                            cell.font = Font(name='Calibri', size=10, bold=True, color=CORES['erro'])

                nome_dif = _safe_filename(f"Totais_Consolidados_{contrato_id}_Diferencas.xlsx")
                caminho_dif = f"{pasta}/{nome_dif}"
                wb_d.save(caminho_dif)
                gerados.append(caminho_dif)

    if gerados:
        msg = "Arquivos gerados:\n- " + "\n- ".join(gerados[:14])
        if len(gerados) > 14:
            msg += f"\n... (+{len(gerados)-14})"
        messagebox.showinfo("Sucesso", msg)
    else:
        messagebox.showwarning("Aviso", "Nada foi gerado (dados vazios ou contratos sem registros).")