# utils/processador_sa.py - Versão refatorada
from __future__ import annotations
import pandas as pd
import tkinter as tk
from tkinter import ttk
import ttkbootstrap as ttkb
from datetime import datetime, date
from typing import Iterable, List, Dict, Any
import sys
from pathlib import Path
import re
from tkinter import filedialog, messagebox
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# ============================= CONFIGURAÇÕES =============================

_ARQ_PAD = Path(__file__).resolve().parents[1] / "data" / "RegistroBoletim.parquet"
ARQUIVO_PADRAO = str(_ARQ_PAD)

IGNORAR_COLS_DEFAULT = {
    "Data", "DATA",
    "Boletim", "BOLETIM",
    "Registro", "REGISTRO",
    "Contrato", "CONTRATO",
    "Arquivo", "ARQUIVO",
    "Fornecedor", "FORNECEDOR",
}

_COLS_SA = ["Funcionário", "Registro", "BOLETIM", "Contrato", "S.A."]
_COLS_SA_UI = ["Funcionário", "Registro", "Boletim", "Contrato", "S.A."]

ALIASES_SA = {
    "Funcionário": {"Funcionário", "FUNCIONÁRIO", "FUNCIONARIO"},
    "Registro": {"Registro", "REGISTRO", "Matrícula", "MATRÍCULA", "MATRICULA"},
    "Boletim": {"Boletim", "BOLETIM"},
    "Contrato": {"Contrato", "CONTRATO"},
    "S.A.": {"S.A.", "SA", "S A", "Sobreaviso", "SOBREAVISO"},
    "Data": {"Data", "DATA"},
}

_COLS_NECS = ["Funcionário", "Registro", "BOLETIM", "Contrato", "S.A.", "DATA"]

# ============================= FUNÇÕES AUXILIARES =============================

def _fmt_data_br(x) -> str:
    """Retorna a data em dd/mm/aaaa, aceitando date/datetime/Timestamp/str."""
    if x is None or x == "":
        return ""
    if isinstance(x, (pd.Timestamp, datetime, date)):
        return pd.Timestamp(x).strftime("%d/%m/%Y")
    s = str(x).strip()
    try:
        return datetime.strptime(s, "%d/%m/%Y").strftime("%d/%m/%Y")
    except ValueError:
        ts = pd.to_datetime(s, dayfirst=True, errors="coerce")
        return ts.strftime("%d/%m/%Y") if pd.notna(ts) else ""

def _fmt_num_sa(v):
    """Formata número S.A. com 3 casas decimais"""
    try:
        return f"{float(v):.3f}".replace(".", ",")
    # Sugestão 4: Ser específico nas exceções
    except (ValueError, TypeError, AttributeError):
        return str(v)

def _to_number_series(s: pd.Series) -> pd.Series:
    """Converte strings com milhar/decimal em número."""
    s = s.astype(str).str.strip().str.replace("\u00A0", "", regex=False)
    s = s.str.replace(r"\.(?=\d{3}(?:\D|$))", "", regex=True)
    s = s.str.replace(",", ".", regex=False)
    return pd.to_numeric(s, errors="coerce")

def _to_number_sa(s: pd.Series) -> pd.Series:
    """Converte S.A. para numérico de forma robusta."""
    if pd.api.types.is_numeric_dtype(s):
        return s.fillna(0.0)
    s = s.astype(str).str.strip().str.replace("\u00A0", "", regex=False)
    s = s.str.replace(r"\.(?=\d{3}(?:\D|$))", "", regex=True)
    s = s.str.replace(",", ".", regex=False)
    return pd.to_numeric(s, errors="coerce").fillna(0.0)

def _coerce_sa_one(x) -> float:
    """Converte S.A. para float de forma robusta (aceita HH:MM também)."""
    if x is None:
        return 0.0
    if isinstance(x, (int, float)) and not pd.isna(x):
        return float(x)

    s = str(x).strip()
    if not s:
        return 0.0
    s = s.replace("\xa0", " ").strip()

    # HH:MM
    m = re.fullmatch(r"\s*(\d{1,3})\s*:\s*(\d{1,2})\s*", s)
    if m:
        hh = int(m.group(1))
        mm = int(m.group(2))
        return float(hh) + (mm / 60.0)

    s = re.sub(r"(?<=\d)[\.\s](?=\d{3}(\D|$))", "", s)
    s = s.replace(",", ".")
    s = re.sub(r"[^0-9\.\-]", "", s)
    try:
        return float(s)
    # Sugestão 4: Ser específico nas exceções
    except (ValueError, TypeError):
        return 0.0

def _coerce_sa(col: pd.Series) -> pd.Series:
    """Aplica a conversão robusta elemento a elemento."""
    return col.apply(_coerce_sa_one)

def _coerce_numeric(df: pd.DataFrame, numeric_only_cols: list[str] | None = None) -> pd.DataFrame:
    """Converte colunas para numéricas (quando possível) sem quebrar strings."""
    work = df.copy()
    cols = numeric_only_cols or [c for c in work.columns if c != "Funcionário"]
    for c in cols:
        if c in work.columns and c != "Funcionário":
            s = work[c]
            work[c] = _to_number_series(s) if s.dtype == object else pd.to_numeric(s, errors="coerce")
    return work

def _normalize_min_cols(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza nomes mínimos usados pelo relatório de S.A."""
    work = df.copy()
    rename_map = {}
    if "DATA" in work.columns and "Data" not in work.columns:
        rename_map["DATA"] = "Data"
    if "BOLETIM" in work.columns and "Boletim" not in work.columns:
        rename_map["BOLETIM"] = "Boletim"
    if "S.A" in work.columns and "S.A." not in work.columns:
        rename_map["S.A"] = "S.A."
    if rename_map:
        work = work.rename(columns=rename_map)
    return work

def _normalize_cols_sa(df: pd.DataFrame) -> pd.DataFrame:
    """Normaliza nomes de colunas usando aliases."""
    if df is None or df.empty:
        return df
    rename_map = {}
    cols = set(df.columns)
    for target, aliases in ALIASES_SA.items():
        for a in list(aliases):
            if a in cols:
                rename_map[a] = target
        for a in list(aliases):
            au = str(a).upper()
            if au in cols:
                rename_map[au] = target
    return df.rename(columns=rename_map).copy()

def _ensure_cols(df: pd.DataFrame, cols: Iterable[str]) -> pd.DataFrame:
    """Garante que as colunas existem no DataFrame."""
    w = df.copy()
    for c in cols:
        if c not in w.columns:
            w[c] = pd.NA
    return w

def _extrair_contratos_unicos(df: pd.DataFrame) -> list:
    """Extrai lista de contratos únicos do DataFrame"""
    if 'Contrato' not in df.columns:
        return []
    contratos = df['Contrato'].dropna().unique()
    contratos = [c for c in contratos if str(c).strip() and str(c).strip() != '']
    return sorted(str(c) for c in contratos)

# ============================= FUNÇÕES DE CABEÇALHO =============================

def _criar_cabecalho_pdf(data_ini, data_fim, contratos=None, titulo="Relatório de Sobreavisos"):
    """Cria elementos de cabeçalho formatados para PDF"""
    styles = getSampleStyleSheet()
    
    titulo_style = ParagraphStyle(
        'TituloCustom',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2F5597'),
        spaceAfter=12,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    info_style = ParagraphStyle(
        'InfoCustom',
        parent=styles['Normal'],
        fontSize=10,
        textColor=colors.HexColor('#333333'),
        spaceAfter=6,
        alignment=TA_LEFT
    )
    
    elementos = []
    elementos.append(Paragraph(titulo, titulo_style))
    elementos.append(Spacer(1, 0.3*cm))
    
    info_lines = []
    
    if data_ini and data_fim:
        periodo = f"<b>Período:</b> {_fmt_data_br(data_ini)} a {_fmt_data_br(data_fim)}"
        info_lines.append(periodo)
    
    if contratos:
        if isinstance(contratos, (list, set)):
            contratos_str = ", ".join(sorted(str(c) for c in contratos if c))
        else:
            contratos_str = str(contratos)
        if contratos_str:
            info_lines.append(f"<b>Contratos:</b> {contratos_str}")
    
    info_lines.append(f"<b>Gerado em:</b> {datetime.now().strftime('%d/%m/%Y às %H:%M')}")
    
    for line in info_lines:
        elementos.append(Paragraph(line, info_style))
    
    elementos.append(Spacer(1, 0.5*cm))
    
    return elementos

def _criar_cabecalho_excel(ws, data_ini, data_fim, contratos=None, titulo="Relatório de Sobreavisos"):
    """Cria cabeçalho formatado para Excel"""
    titulo_font = Font(name='Calibri', size=14, bold=True, color='2F5597')
    titulo_fill = PatternFill(start_color='E8F0F8', end_color='E8F0F8', fill_type='solid')
    
    info_font = Font(name='Calibri', size=10)
    info_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    
    row = 1
    
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = titulo
    cell.font = titulo_font
    cell.fill = titulo_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    row += 1
    
    row += 1
    
    if data_ini and data_fim:
        ws.merge_cells(f'A{row}:E{row}')
        cell = ws[f'A{row}']
        cell.value = f"Período: {_fmt_data_br(data_ini)} a {_fmt_data_br(data_fim)}"
        cell.font = info_font
        cell.fill = info_fill
        cell.alignment = Alignment(horizontal='left', vertical='center')
        row += 1
    
    if contratos:
        if isinstance(contratos, (list, set)):
            contratos_str = ", ".join(sorted(str(c) for c in contratos if c))
        else:
            contratos_str = str(contratos)
        if contratos_str:
            ws.merge_cells(f'A{row}:E{row}')
            cell = ws[f'A{row}']
            cell.value = f"Contratos: {contratos_str}"
            cell.font = info_font
            cell.fill = info_fill
            cell.alignment = Alignment(horizontal='left', vertical='center')
            row += 1
    
    ws.merge_cells(f'A{row}:E{row}')
    cell = ws[f'A{row}']
    cell.value = f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M')}"
    cell.font = info_font
    cell.fill = info_fill
    cell.alignment = Alignment(horizontal='left', vertical='center')
    row += 1
    
    row += 2
    
    return row

# ============================= FUNÇÕES DE CARREGAMENTO =============================

def _carregar_parquet(caminho: str) -> pd.DataFrame:
    """Carrega o parquet e retorna DataFrame."""
    try:
        df = pd.read_parquet(caminho)
        return df
    except ImportError as e:
        msg = str(e).lower()
        print(
            "[ERRO] Suporte a Parquet não encontrado.\n"
            "Instale um motor: 'pip install pyarrow' (recomendado) ou 'pip install fastparquet'.",
            file=sys.stderr,
        )
        raise
    except FileNotFoundError:
        print(f"[ERRO] Arquivo não encontrado: {caminho}", file=sys.stderr)
        raise
    except Exception as e:
        print(f"[ERRO] Falha ao ler parquet '{caminho}': {e}", file=sys.stderr)
        raise

def carregar_boletim_parquet(caminho: str = ARQUIVO_PADRAO) -> pd.DataFrame:
    return pd.read_parquet(caminho)

# ============================= FUNÇÕES DE RELATÓRIO =============================

# +++ Sugestão 1: Centralizar lógica de subtotal +++
def _adicionar_subtotais_registro(
    df_agrupado: pd.DataFrame, 
    col_agrupadora: str, 
    col_boletim: str
) -> pd.DataFrame:
    """
    Itera sobre um DF agrupado por 'col_agrupadora' e insere linhas de subtotal
    específicas para o relatório por registro.
    """
    partes = []
    colunas_subtotal = [col for col in df_agrupado.columns if col != col_agrupadora]
    
    for nome, bloco in df_agrupado.groupby(col_agrupadora, sort=False):
        partes.append(bloco)
        # Adiciona subtotal apenas se houver mais de 1 boletim
        if bloco[col_boletim].nunique() > 1:
            subtotal_row = {
                col_agrupadora: f"Subtotal {nome}",
                "S.A.": float(bloco["S.A."].sum())
            }
            # Preenche as outras colunas com ""
            for col in colunas_subtotal:
                if col not in subtotal_row:
                    subtotal_row[col] = ""
                    
            partes.append(pd.DataFrame([subtotal_row], columns=df_agrupado.columns))

    if not partes:
        return df_agrupado

    return pd.concat(partes, ignore_index=True)
# +++ Fim da Sugestão 1 +++


def _forcar_tipos_sa(df: pd.DataFrame) -> pd.DataFrame:
    """Garante colunas e tipos corretos para somatório de S.A."""
    work = df.copy()

    if "Boletim" in work.columns and "BOLETIM" not in work.columns:
        work["BOLETIM"] = work["Boletim"]

    for c in ["Funcionário", "Registro", "BOLETIM", "Contrato"]:
        if c not in work.columns:
            work[c] = ""
        work[c] = work[c].astype(str).fillna("")

    if "S.A." not in work.columns:
        work["S.A."] = 0.0
    work["S.A."] = _to_number_sa(work["S.A."])

    return work[_COLS_SA].copy()

def relatorio_sa_por_registro(df: pd.DataFrame) -> pd.DataFrame:
    """Retorna DataFrame com: Funcionário | Registro | BOLETIM | Contrato | S.A."""
    if df is None or df.empty:
        return pd.DataFrame(columns=_COLS_SA)

    base = _forcar_tipos_sa(df)

    grouped = (
        base.groupby(["Funcionário", "Registro", "BOLETIM", "Contrato"], as_index=False)["S.A."]
        .sum()
        .sort_values(["Funcionário", "BOLETIM", "Registro", "Contrato"], kind="stable")
        .reset_index(drop=True)
    )

    # Sugestão 1: Substitui o loop de subtotal
    saida = _adicionar_subtotais_registro(grouped, "Funcionário", "BOLETIM")
    
    return saida.rename(columns={"BOLETIM": "Boletim"})[_COLS_SA_UI]


def relatorio_sa_por_registro_periodo(
    df: pd.DataFrame,
    data_ini,
    data_fim,
) -> pd.DataFrame:
    """Soma S.A. por (Funcionário, Registro, BOLETIM, Contrato) apenas no período."""
    cols_saida = ["Funcionário", "Registro", "Boletim", "Contrato", "S.A."]
    if df is None or df.empty:
        return pd.DataFrame(columns=cols_saida)

    work = df.copy()

    rename_in = {}
    if "Boletim" in work.columns and "BOLETIM" not in work.columns:
        rename_in["Boletim"] = "BOLETIM"
    if "Data" in work.columns and "DATA" not in work.columns:
        rename_in["Data"] = "DATA"
    if rename_in:
        work = work.rename(columns=rename_in)

    for c in ["Funcionário", "Registro", "BOLETIM", "Contrato", "S.A.", "DATA"]:
        if c not in work.columns:
            work[c] = pd.NA

    work["DATA"] = pd.to_datetime(work["DATA"], dayfirst=True, errors="coerce")

    if isinstance(data_ini, str):
        data_ini = pd.to_datetime(data_ini, dayfirst=True, errors="coerce")
    if isinstance(data_fim, str):
        data_fim = pd.to_datetime(data_fim, dayfirst=True, errors="coerce")

    if data_ini is None or pd.isna(data_ini):
        data_ini = work["DATA"].min()
    if data_fim is None or pd.isna(data_fim):
        data_fim = work["DATA"].max()

    work = work[(work["DATA"] >= data_ini) & (work["DATA"] <= data_fim)].copy()
    if work.empty:
        return pd.DataFrame(columns=cols_saida)

    work["S.A."] = _coerce_sa(work["S.A."])
    for c in ["Funcionário", "Registro", "BOLETIM", "Contrato"]:
        work[c] = work[c].astype(str).fillna("")

    base = (
        work.groupby(["Funcionário", "Registro", "BOLETIM", "Contrato"], as_index=False)["S.A."]
            .sum()
            .sort_values(["Funcionário", "BOLETIM", "Registro", "Contrato"], kind="stable")
            .reset_index(drop=True)
    )

    # Sugestão 1: Substitui o loop de subtotal
    saida = _adicionar_subtotais_registro(base, "Funcionário", "BOLETIM")

    saida = saida.rename(columns={"BOLETIM": "Boletim"})
    return saida[cols_saida]

def relatorio_sa_por_func_contrato_boletim(
    df: pd.DataFrame,
    *,
    ordenar_por=("Funcionário", "Contrato", "Boletim"),
    incluir_zeros: bool = False,
) -> pd.DataFrame:
    """S.A. por Funcionário → Contrato → Boletim"""
    if df is None or df.empty:
        return pd.DataFrame(columns=["Funcionário", "Contrato", "Boletim", "S.A."])

    work = _normalize_min_cols(df)

    for c in ("Funcionário", "Contrato", "Boletim"):
        if c not in work.columns:
            work[c] = ""

    has_data = "Data" in work.columns
    if has_data:
        work["Data"] = pd.to_datetime(work["Data"], dayfirst=True, errors="coerce").dt.date

    if "S.A." not in work.columns:
        work["S.A."] = 0.0
    work["S.A."] = _to_number_series(work["S.A."])

    subset_dup = ["Funcionário", "Contrato", "Boletim", "S.A."]
    if has_data:
        subset_dup.append("Data")
    work = work.drop_duplicates(subset=subset_dup, keep="last")

    if has_data:
        daily = (
            work.groupby(["Funcionário", "Contrato", "Boletim", "Data"], dropna=False)["S.A."]
            .sum()
            .reset_index()
        )
    else:
        daily = work[["Funcionário", "Contrato", "Boletim", "S.A."]].copy()

    base = (
        daily.groupby(["Funcionário", "Contrato", "Boletim"], dropna=False)["S.A."]
        .sum()
        .reset_index()
    )

    if not incluir_zeros:
        base = base[base["S.A."].fillna(0) != 0]

    if ordenar_por:
        base = base.sort_values(list(ordenar_por), kind="stable", ignore_index=True)

    partes = []
    # Nota: Este loop de subtotal é DIFERENTE, por isso não foi refatorado
    for nome, bloco in base.groupby("Funcionário", sort=False):
        bloco = bloco.sort_values(["Contrato", "Boletim"], kind="stable")
        partes.append(bloco)
        if len(bloco) > 1:
            partes.append(pd.DataFrame([{
                "Funcionário": f"*{nome}*",
                "Contrato": "",
                "Boletim": "",
                "S.A.": bloco["S.A."].sum()
            }]))

    saida = pd.concat(partes, ignore_index=True) if partes else base
    return saida[["Funcionário", "Contrato", "Boletim", "S.A."]]

def totalizar_sa_func_contrato_boletim(df: pd.DataFrame, **kwargs) -> pd.DataFrame:
    """Alias para relatorio_sa_por_func_contrato_boletim"""
    return relatorio_sa_por_func_contrato_boletim(df, **kwargs)

def montar_sa_por_registro(
    df_boletim: pd.DataFrame,
    *,
    data_ini,
    data_fim,
    incluir_subtotal: bool = True,
) -> pd.DataFrame:
    """Retorna DataFrame com: Funcionário | Registro | Boletim | Contrato | S.A."""
    cols_saida = ["Funcionário", "Registro", "Boletim", "Contrato", "S.A."]
    if df_boletim is None or df_boletim.empty:
        return pd.DataFrame(columns=cols_saida)

    work = _normalize_cols_sa(df_boletim).copy()

    for c in ("Funcionário", "Registro", "Boletim", "Contrato", "S.A."):
        if c not in work.columns:
            work[c] = "" if c != "S.A." else 0.0

    if "Data" in work.columns:
        work["Data"] = pd.to_datetime(work["Data"], dayfirst=True, errors="coerce")

    def _to_dt(s):
        import pandas as _pd
        if isinstance(s, str):
            return _pd.to_datetime(s, dayfirst=True, errors="coerce")
        return _pd.to_datetime(s, errors="coerce")
    
    di = _to_dt(data_ini)
    df = _to_dt(data_fim)
    if "Data" in work.columns and pd.notna(di) and pd.notna(df):
        work = work[(work["Data"] >= di) & (work["Data"] <= df)]

    work["S.A."] = _to_number_sa(work["S.A."])

    keys = ["Funcionário", "Registro", "Boletim", "Contrato"]
    base = work.groupby(keys, dropna=False)["S.A."].sum().reset_index()
    base = base.sort_values(keys, kind="stable", ignore_index=True)

    if not incluir_subtotal:
        return base[cols_saida]

    # Sugestão 1: Substitui o loop de subtotal
    saida = _adicionar_subtotais_registro(base, "Funcionário", "Boletim")
    
    return saida[cols_saida]

def gerar_df_sa_por_registro_periodo(
    caminho: str = ARQUIVO_PADRAO, data_ini=None, data_fim=None
) -> pd.DataFrame:
    """Lê o parquet e retorna o DF de S.A. agregado por registro no período."""
    df = carregar_boletim_parquet(caminho)
    # Nota: Esta função chama 'relatorio_sa_por_registro_periodo', que foi
    # mantida por consistência, embora 'montar_sa_por_registro' seja similar.
    return relatorio_sa_por_registro_periodo(df, data_ini, data_fim)

def totalizar_por_funcionario(
    df: pd.DataFrame,
    *,
    coluna_func: str = "Funcionário",
    ignorar_cols: set[str] | None = None,
    ordenar_por_nome: bool = True,
) -> pd.DataFrame:
    """Soma todas as colunas numéricas por Funcionário."""
    if df is None or df.empty:
        return pd.DataFrame(columns=[coluna_func])

    ignorar = set(ignorar_cols or IGNORAR_COLS_DEFAULT)
    if coluna_func not in df.columns:
        raise KeyError(f"Coluna '{coluna_func}' não encontrada no DataFrame.")

    candidatas = [c for c in df.columns if c not in ignorar and c != coluna_func]
    work = _coerce_numeric(df[[coluna_func] + candidatas], numeric_only_cols=candidatas)

    num_cols = [c for c in candidatas if pd.api.types.is_numeric_dtype(work[c])]
    if not num_cols:
        return pd.DataFrame(columns=[coluna_func])

    agrupado = work.groupby(coluna_func, dropna=False, sort=ordenar_por_nome)[num_cols].sum().reset_index()

    if ordenar_por_nome:
        agrupado = agrupado.sort_values(coluna_func, kind="stable", ignore_index=True)

    total = agrupado[num_cols].sum(numeric_only=True)
    total_row = pd.DataFrame([{coluna_func: "TOTAL GERAL", **total.to_dict()}])
    return pd.concat([agrupado, total_row], ignore_index=True)

# ============================= CONSOLE =============================

# --- Sugestão 2: Funções removidas ---
# As funções `_validar_colunas`, `_forcar_tipos`, `_agrupar_por_registro` e
# `_montar_linhas_com_subtotais` foram removidas pois eram uma
# reimplementação da lógica já existente em `relatorio_sa_por_registro`.
# --- Fim da Sugestão 2 ---

def _formatar_e_imprimir(linhas: List[Dict[str, Any]]):
    """Formata e imprime as linhas do relatório."""
    if not linhas:
        print("Nenhum dado para exibir.")
        return
    
    print("\n" + "="*80)
    print("RELATÓRIO DE SOBREAVISOS POR REGISTRO")
    print("="*80)
    print(f"{'Funcionário':<40} {'Registro':<12} {'Boletim':<12} {'Contrato':<20} {'S.A.':>10}")
    print("-"*80)
    
    total_sa_geral = 0.0
    
    for linha in linhas:
        func = linha["Funcionário"]
        reg = linha["Registro"]
        # 'relatorio_sa_por_registro' retorna 'Boletim'
        bol = linha["Boletim"] 
        cont = linha["Contrato"]
        sa = linha["S.A."]
        
        try:
            sa_float = float(sa)
            if not str(func).startswith("Subtotal"):
                total_sa_geral += sa_float
            print(f"{func:<40} {reg:<12} {bol:<12} {cont:<20} {sa_float:>10.3f}")
        except (ValueError, TypeError):
            print(f"{func:<40} {reg:<12} {bol:<12} {cont:<20} {str(sa):>10}")
    
    print("="*80)
    print(f"{'TOTAL GERAL':<70} {total_sa_geral:>10.3f}")
    print("="*80 + "\n")

def gerar_relatorio_console(caminho_arquivo: str = ARQUIVO_PADRAO):
    """Gera relatório de console com dados do parquet."""
    # Sugestão 2: Simplificado para usar a função de relatório principal
    try:
        df = _carregar_parquet(caminho_arquivo)
    except Exception as e:
        print(f"[ERRO] Falha ao carregar dados: {e}", file=sys.stderr)
        return
        
    # 1. Chame a função de relatório principal (ela já trata tipos e subtotais)
    df_relatorio = relatorio_sa_por_registro(df)
    
    # 2. Converta o DataFrame resultante para uma lista de dicts
    # (relatorio_sa_por_registro já retorna as colunas corretas)
    linhas = df_relatorio.to_dict('records')

    # 3. Imprima
    _formatar_e_imprimir(linhas)


# ============================= LÓGICA DE EXPORTAÇÃO (SEPARADA) =============================

# +++ Sugestão 3: Mover lógica de exportação para fora da UI +++

def exportar_sa_para_excel(
    df_view: pd.DataFrame, 
    caminho_salvar: str, 
    data_ini=None, 
    data_fim=None, 
    contratos=None
):
    """Salva o DataFrame do relatório de S.A. em um arquivo Excel formatado."""
    # A lógica de 'try/except' deve ficar na UI, que pode mostrar messagebox
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sobreavisos"

    # Criar cabeçalho
    linha_inicial = _criar_cabecalho_excel(
        ws, data_ini, data_fim, contratos,
        "Relatório de Sobreavisos por Registro"
    )

    # Estilos para cabeçalho da tabela
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2F5597', end_color='2F5597', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Cabeçalhos das colunas
    headers = ["Funcionário", "Registro", "Boletim", "Contrato", "S.A."]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=linha_inicial, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border

    # Dados
    row_idx = linha_inicial + 1
    for _, row_data in df_view.iterrows():
        for col_idx, col_name in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            valor = row_data.get(col_name, "")
            
            if col_name == "S.A.":
                try:
                    cell.value = float(valor)
                except (ValueError, TypeError):
                    cell.value = 0.0
                cell.number_format = '#,##0.000'
                cell.alignment = Alignment(horizontal='right')
            else:
                cell.value = str(valor)
                cell.alignment = Alignment(horizontal='left')
            
            cell.border = border
            
            # Destaque para subtotais
            if str(row_data.get("Funcionário", "")).startswith("Subtotal"):
                cell.font = Font(bold=True, color='0000FF')
                cell.fill = PatternFill(start_color='E8F0FF', end_color='E8F0FF', fill_type='solid')
        
        row_idx += 1

    # Linha de total
    row_idx += 1
    total_cell = ws.cell(row=row_idx, column=1)
    total_cell.value = "TOTAL GERAL"
    total_cell.font = Font(bold=True, size=11)
    total_cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
    
    total_sa = float(pd.to_numeric(df_view["S.A."], errors="coerce").fillna(0).sum())
    
    total_sa_cell = ws.cell(row=row_idx, column=5)
    total_sa_cell.value = total_sa
    total_sa_cell.number_format = '#,##0.000'
    total_sa_cell.font = Font(bold=True, size=11)
    total_sa_cell.fill = PatternFill(start_color='FFE699', end_color='FFE699', fill_type='solid')
    total_sa_cell.alignment = Alignment(horizontal='right')

    # Ajustar larguras
    col_widths = {"Funcionário": 35, "Registro": 15, "Boletim": 15, "Contrato": 25, "S.A.": 12}
    for col_idx, (col_name, width) in enumerate(col_widths.items(), 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    wb.save(caminho_salvar)
    
def exportar_sa_para_pdf(
    df_view: pd.DataFrame, 
    caminho_salvar: str, 
    data_ini=None, 
    data_fim=None, 
    contratos=None
):
    """Salva o DataFrame do relatório de S.A. em um arquivo PDF formatado."""
    doc = SimpleDocTemplate(
        caminho_salvar, pagesize=landscape(A4),
        rightMargin=20, leftMargin=20, topMargin=20, bottomMargin=20
    )
    
    elementos = []
    
    # Adicionar cabeçalho
    elementos.extend(_criar_cabecalho_pdf(
        data_ini, data_fim, contratos,
        "Relatório de Sobreavisos por Registro"
    ))

    # Tabela de dados
    headers = ["Funcionário", "Registro", "Boletim", "Contrato", "S.A."]
    data = [headers]
    
    for _, row in df_view.iterrows():
        try:
            sa_val = f"{float(row.get('S.A.', 0.0)):.3f}"
        except (ValueError, TypeError):
            sa_val = str(row.get('S.A.', '0,000'))
            
        data.append([
            str(row.get("Funcionário", "")),
            str(row.get("Registro", "")),
            str(row.get("Boletim", "")),
            str(row.get("Contrato", "")),
            sa_val,
        ])

    # Linha de total
    total_sa = float(pd.to_numeric(df_view["S.A."], errors="coerce").fillna(0).sum())
    data.append([
        "TOTAL GERAL", "", "", "", f"{total_sa:.3f}"
    ])

    table = Table(data, repeatRows=1)
    
    # Estilos da tabela
    table_style = [
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F5597")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 11),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("ALIGN", (-1, 1), (-1, -1), "RIGHT"), # Alinha S.A. à direita
        ("ALIGN", (0, 1), (-2, -1), "LEFT"), # Alinha texto à esquerda
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), 
         [colors.whitesmoke, colors.HexColor("#F8F8F8")]),
        ("BOTTOMPADDING", (0, 0), (-1, 0), 10),
        ("TOPPADDING", (0, 0), (-1, 0), 10),
        # Linha de total
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFE699")),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("FONTSIZE", (0, -1), (-1, -1), 11),
    ]
    
    # Destaca subtotais
    for i, row in enumerate(df_view.to_dict('records'), 1):
        if str(row.get('Funcionário', '')).startswith('Subtotal'):
            table_style.extend([
                ("TEXTCOLOR", (0, i), (-1, i), colors.blue),
                ("FONTNAME", (0, i), (-1, i), "Helvetica-Bold"),
                ("BACKGROUND", (0, i), (-1, i), colors.HexColor("#E8F0FF")),
            ])
    
    table.setStyle(TableStyle(table_style))
    elementos.append(table)
    
    # Resumo
    elementos.append(Spacer(1, 0.6*cm))
    styles = getSampleStyleSheet()
    resumo_style = ParagraphStyle(
        'ResumoCustom',
        parent=styles['Normal'],
        fontSize=11,
        textColor=colors.HexColor('#2F5597'),
        fontName='Helvetica-Bold'
    )
    
    total_registros = len([r for r in df_view.to_dict('records') 
                          if not str(r.get('Funcionário', '')).startswith('Subtotal')])
    
    elementos.append(Paragraph(
        f"Total de Sobreaviso (S.A.): {total_sa:.3f} horas | "
        f"Total de Registros: {total_registros}",
        resumo_style
    ))
    
    doc.build(elementos)

# +++ Fim da Sugestão 3 +++


# ============================= INTERFACE GRÁFICA =============================

class RelatorioSADialog(ttkb.Toplevel):
    """Janela de diálogo para exibir relatório de S.A. com cabeçalhos melhorados"""
    
    def __init__(self, parent, df: pd.DataFrame, *, data_ini=None, data_fim=None):
        super().__init__(parent)
        self.title("Relatório de Sobreavisos por Registro")
        self.geometry("1100x680")
        self.minsize(900, 500)
        #self.transient(parent)
        self.grab_set()

        # Guarda dados
        if isinstance(df, pd.DataFrame):
            self.df_view = df.copy()
        else:
            self.df_view = pd.DataFrame()

        self.data_ini = data_ini
        self.data_fim = data_fim
        self.contratos = _extrair_contratos_unicos(self.df_view)

        # Normaliza colunas
        if "BOLETIM" in self.df_view.columns and "Boletim" not in self.df_view.columns:
            self.df_view = self.df_view.rename(columns={"BOLETIM": "Boletim"})
        
        for c in _COLS_SA_UI:
            if c not in self.df_view.columns:
                self.df_view[c] = ""
        
        # Força numérico para S.A.
        self.df_view["S.A."] = pd.to_numeric(
            self.df_view["S.A."], errors="coerce"
        ).fillna(0.0)

        self._criar_interface()

    def _criar_interface(self):
        # Frame de cabeçalho com informações do relatório
        frame_header = ttkb.Frame(self, padding=(15, 12), bootstyle="light")
        frame_header.pack(fill="x")

        # Título
        titulo = ttkb.Label(
            frame_header,
            text="📊 Relatório de Sobreavisos por Registro",
            font=("Segoe UI", 12, "bold"),
            bootstyle="primary"
        )
        titulo.pack(anchor="w", pady=(0, 8))

        # Frame para informações
        info_frame = ttkb.Frame(frame_header)
        info_frame.pack(fill="x")

        # Período
        if self.data_ini and self.data_fim:
            periodo_txt = f"📅 Período: {_fmt_data_br(self.data_ini)} a {_fmt_data_br(self.data_fim)}"
            ttkb.Label(info_frame, text=periodo_txt, bootstyle="secondary").pack(
                side="left", padx=(0, 20)
            )

        # Contratos
        if self.contratos:
            contratos_txt = f"📋 Contratos: {', '.join(self.contratos)}"
            ttkb.Label(info_frame, text=contratos_txt, bootstyle="secondary").pack(
                side="left", padx=(0, 20)
            )

        # Data de geração
        geracao_txt = f"🕐 Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ttkb.Label(info_frame, text=geracao_txt, bootstyle="secondary").pack(side="right")

        # Separador
        ttk.Separator(self, orient="horizontal").pack(fill="x", pady=8)

        # Frame da tabela
        frame_tbl = ttkb.Frame(self, padding=(15, 0))
        frame_tbl.pack(fill="both", expand=True)

        self.tree = ttk.Treeview(frame_tbl, columns=_COLS_SA_UI, show="headings")
        vsb = ttkb.Scrollbar(frame_tbl, orient="vertical", command=self.tree.yview)
        hsb = ttkb.Scrollbar(frame_tbl, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        
        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        
        frame_tbl.rowconfigure(0, weight=1)
        frame_tbl.columnconfigure(0, weight=1)

        for c in _COLS_SA_UI:
            self.tree.heading(c, text=c, command=lambda col=c: self._ordenar_por(col, False))

        self._popular(self.df_view)

        # Separador
        ttk.Separator(self, orient="horizontal").pack(fill="x", pady=8)

        # Rodapé com resumo e botões
        rodape = ttkb.Frame(self, padding=(15, 10), bootstyle="light")
        rodape.pack(fill="x")

        # Frame esquerdo - informações
        info_left = ttkb.Frame(rodape)
        info_left.pack(side="left", fill="x", expand=True)

        total_sa = self._calc_total()
        total_registros = len([r for r in self.df_view.to_dict('records') 
                              if not str(r.get('Funcionário', '')).startswith('Subtotal')])

        self.lbl_total = ttkb.Label(
            info_left,
            text=f"💰 Total S.A.: {total_sa:.3f} h  |  📝 Registros: {total_registros}",
            bootstyle="info",
            font=("Segoe UI", 10, "bold"),
        )
        self.lbl_total.pack(side="left")

        # Frame direito - botões
        btn_frame = ttkb.Frame(rodape)
        btn_frame.pack(side="right")

        ttkb.Button(
            btn_frame, text="📄 Exportar PDF", bootstyle="danger-outline",
            command=self._exportar_pdf_ui # Sugestão 3: Chama o wrapper da UI
        ).pack(side="right", padx=4)
        
        ttkb.Button(
            btn_frame, text="📊 Exportar Excel", bootstyle="success-outline",
            command=self._exportar_excel_ui # Sugestão 3: Chama o wrapper da UI
        ).pack(side="right", padx=4)
        
        ttkb.Button(
            btn_frame, text="✖ Fechar", command=self.destroy,
            bootstyle="secondary-outline"
        ).pack(side="right", padx=4)

    def _popular(self, df: pd.DataFrame):
        """Popula a TreeView com os dados"""
        for iid in self.tree.get_children():
            self.tree.delete(iid)

        for _, row in df.iterrows():
            vals = [
                row["Funcionário"], row["Registro"], row["Boletim"], 
                row["Contrato"], _fmt_num_sa(row["S.A."]),
            ]
            item = self.tree.insert("", "end", values=vals)
            if str(row["Funcionário"]).startswith("Subtotal "):
                self.tree.item(item, tags=("subtotal",))
        
        self.tree.tag_configure("subtotal", foreground="blue", font=("Segoe UI", 9, "bold"))
        self._ajustar_larguras()

    def _ajustar_larguras(self):
        """Ajusta largura das colunas automaticamente"""
        minw = {"Funcionário": 260, "Registro": 90, "Boletim": 110, "Contrato": 140, "S.A.": 100}
        for c in _COLS_SA_UI:
            maxlen = len(c)
            for k, iid in enumerate(self.tree.get_children()):
                if k > 500:
                    break
                txt = str(self.tree.set(iid, c))
                maxlen = max(maxlen, len(txt))
            width = max(minw.get(c, 80), min(520, int(maxlen * 7.2) + 16))
            self.tree.column(c, width=width, anchor="e" if c == "S.A." else "w")

    def _ordenar_por(self, col: str, reverso: bool):
        """Ordena a TreeView por coluna"""
        dados = [(self.tree.set(k, col), k) for k in self.tree.get_children("")]
        if col == "S.A.":
            def _key(x):
                try:
                    return float(str(x[0]).replace(".", "").replace(",", "."))
                # Sugestão 4: Ser específico
                except (ValueError, TypeError):
                    return 0.0
        else:
            def _key(x):
                return str(x[0]).casefold()
        dados.sort(key=_key, reverse=reverso)
        for idx, (_, k) in enumerate(dados):
            self.tree.move(k, "", idx)
        self.tree.heading(col, command=lambda: self._ordenar_por(col, not reverso))

    def _calc_total(self) -> float:
        """Calcula o total de S.A."""
        return float(pd.to_numeric(self.df_view["S.A."], errors="coerce").fillna(0).sum())

    # --- Sugestão 3: Funções de UI que chamam a lógica de exportação ---
    def _exportar_excel_ui(self):
        """Wrapper da UI para exportar Excel."""
        if self.df_view.empty:
            messagebox.showwarning("Aviso", "Não há dados para exportar.")
            return
        
        caminho = filedialog.asksaveasfilename(
            title="Salvar como Excel",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx")],
        )
        if not caminho:
            return
        
        try:
            # Chama a lógica de exportação separada
            exportar_sa_para_excel(
                self.df_view, caminho, self.data_ini, self.data_fim, self.contratos
            )
            messagebox.showinfo("Sucesso", f"Excel salvo em:\n{caminho}")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar Excel:\n{str(e)}")

    def _exportar_pdf_ui(self):
        """Wrapper da UI para exportar PDF."""
        if self.df_view.empty:
            messagebox.showwarning("Aviso", "Não há dados para exportar.")
            return
        
        caminho = filedialog.asksaveasfilename(
            title="Salvar como PDF",
            defaultextension=".pdf",
            filetypes=[("PDF", "*.pdf")],
        )
        if not caminho:
            return
        
        try:
            # Chama a lógica de exportação separada
            exportar_sa_para_pdf(
                self.df_view, caminho, self.data_ini, self.data_fim, self.contratos
            )
            messagebox.showinfo("Sucesso", f"PDF salvo em:\n{caminho}")
        except Exception as e:
            messagebox.showerror("Erro", f"Falha ao exportar PDF:\n{str(e)}")
