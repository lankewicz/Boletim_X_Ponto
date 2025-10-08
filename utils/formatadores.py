# utils/formatadores.py
# ------------------------------------------------------------
# Conversões e formatações de horas (decimal ↔ HH:MM),
# parsing robusto de números/tempos, soma preservando formato
# e utilitários de estilização para tksheet.
#
# Este módulo centraliza lógica que hoje aparece duplicada em:
# - interface/app.py (formatar_tempo)
# - interface/exportador.py (_to_float, _is_hhmm_string, _minutes_to_hhmm, _sum_column_like_source)
# - extrator/leitor_ponto.py (tempo_para_decimal)
#
# Objetivo: evitar duplicação e manter consistência visual/numérica.
#
# Autor: Valdinei Lankewicz
# Criado: 2025-08-13

from __future__ import annotations

from datetime import date, datetime
from typing import Any, Iterable
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

# utils/formatadores.py
from dataclasses import dataclass
from typing import Dict, Iterable, List, Optional, Sequence, Tuple, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

# —— Estilos base ——
THIN = Side(style="thin", color="000000")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# Formatos pt-BR
NUM_FMT_2C = "#.##0,00"
NUM_FMT_INT = "#.##0"
NUM_FMT_MONEY = '"R$" * #.##0,00'


# ============================================================
# Núcleo: conversões HH:MM ↔ decimal
# ============================================================

def _parse_num(val):
    """
    Converte '1.234,56' ou '1234.56' em float. Retorna None se não der.
    """
    try:
        if isinstance(val, (int, float)):
            return float(val)
        s = str(val).strip()
        if not s or s.upper() == "TOTAIS:":
            return None
        # remove separador de milhar e normaliza decimal
        # casos: "1.234,56" -> "1234,56" -> "1234.56"
        if "," in s and "." in s:
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", ".")
        return float(s)
    except Exception:
        return None
    
def hhmm_para_decimal(s: Any) -> float:
    """
    Converte 'HH:MM' (aceita sinal) em horas decimais (float).
    Também aceita números decimais com vírgula/ponto.
    Casos inválidos retornam 0.0.
    """
    if s is None:
        return 0.0
    st = str(s).strip()
    if not st or st == "-":
        return 0.0

    # normaliza separador decimal eventual
    st_norm = st.replace(",", ".", 1)

    # forma HH:MM (com ou sem sinal)
    if ":" in st_norm:
        neg = st_norm.startswith("-")
        if neg:
            st_norm = st_norm[1:]
        try:
            hh, mm = st_norm.split(":", 1)
            h = int(hh)
            m = int(mm)
            total = h + (m / 60.0)
            return -total if neg else total
        except Exception:
            return 0.0

    # decimal (BR/US)
    try:
        return float(st_norm)
    except Exception:
        # tenta remover separadores de milhar comuns
        try:
            st2 = st.replace(".", "").replace(",", ".", 1)
            return float(st2)
        except Exception:
            return 0.0


def decimal_para_hhmm(horas: Any) -> str:
    """
    Converte horas decimais (float/int/str) em 'H:MM' com sinal quando negativo.
    Usa horas sem zero à esquerda (ex.: 2:05, 12:00, -0:15).
    """
    try:
        h = float(str(horas).replace(",", ".", 1))
    except Exception:
        return str(horas)

    neg = h < 0
    h_abs = abs(h)
    hh = int(h_abs)
    mm = round((h_abs - hh) * 60)

    if mm == 60:  # evita 01:60
        hh += 1
        mm = 0

    s = f"{hh}:{mm:02d}"
    return f"-{s}" if neg and s != "0:00" else (f"-0:{mm:02d}" if neg and hh == 0 else s)


# ============================================================
# Parsing genérico e formatação final de exibição
# ============================================================


def parse_num_ou_tempo(valor: Any) -> float:
    """
    Interpreta strings em formato decimal (BR/US) ou 'HH:MM' (com sinal)
    e retorna horas decimais (float). Inválidos viram 0.0.
    """
    if valor in (None, "", "-"):
        return 0.0
    s = str(valor).strip()
    if ":" in s or (s.startswith("-") and ":" in s[1:]):
        return hhmm_para_decimal(s)

    # decimal BR/US, com possíveis milhares
    try:
        # caso: "1.234,56" → 1234.56
        if s.count(",") == 1 and s.count(".") > 1:
            return float(s.replace(".", "").replace(",", ".", 1))
        return float(s.replace(",", ".", 1))
    except Exception:
        try:
            return float(s.replace(".", "").replace(",", ".", 1))
        except Exception:
            return 0.0


def formatar_tempo(valor: Any, *, modo_decimal: bool = True, separador_decimal: str = ",") -> str:
    """
    Formata para exibição:
      - modo_decimal=True  → '12,34' (ou '12.34' se separador_decimal='.')
      - modo_decimal=False → 'H:MM' (com sinal quando negativo)
    Mantém coerência com entrada decimal/str.
    """
    try:
        v = float(str(valor).replace(",", ".", 1))
    except Exception:
        # se já é um HH:MM ou algo não numérico, devolve como veio
        s = str(valor)
        if ":" in s:
            return s
        return s

    if modo_decimal:
        out = f"{v:.2f}"
        return out.replace(".", separador_decimal)
    else:
        return decimal_para_hhmm(v)


# ============================================================
# Soma “preservando formato dominante” (colunas/series)
# ============================================================


def _eh_hhmm(s: Any) -> bool:
    if s is None:
        return False
    st = str(s).strip()
    if not st:
        return False
    # aceita até 3 dígitos de horas, com sinal opcional
    import re

    return bool(re.match(r"^-?\d{1,3}:\d{2}$", st))


def _hhmm_para_minutos(s: str) -> int:
    neg = s.startswith("-")
    if neg:
        s = s[1:]
    hh, mm = s.split(":", 1)
    total = int(hh) * 60 + int(mm)
    return -total if neg else total


def _minutos_para_hhmm(total_min: int) -> str:
    sign = "-" if total_min < 0 else ""
    t = abs(total_min)
    hh = t // 60
    mm = t % 60
    return f"{sign}{hh}:{mm:02d}"


def somar_preservando_formato(series: Iterable[Any], *, default_decimal: bool = True) -> str:
    """
    Soma uma coleção preservando o formato dominante:
      - Se a maioria dos valores válidos for HH:MM → soma em minutos e retorna 'H:MM' (±).
      - Caso contrário → soma decimal e retorna '0,00' (vírgula como default).
    Colunas com texto retornam "" (sem somatório).
    """
    vals = [v for v in series if str(v).strip() not in ("", "-", "None", "nan")]
    if not vals:
        return "0:00" if not default_decimal else "0,00"

    hhmm_count = sum(1 for v in vals if _eh_hhmm(v))

    # Predomina HH:MM → soma como tempo
    if hhmm_count >= max(1, len(vals) // 2):
        total_min = sum(_hhmm_para_minutos(str(v)) for v in vals if _eh_hhmm(v))
        return _minutos_para_hhmm(total_min)

    # Caso contrário, tenta decimal
    parsed = []
    for v in vals:
        s = str(v).strip()
        try:
            parsed.append(
                float(s.replace(".", "").replace(",", ".", 1))
                if s.count(",") == 1 and s.count(".") > 1
                else float(s.replace(",", ".", 1))
            )
        except Exception:
            # texto no meio → não somar
            return ""
    # retorna com vírgula por padrão
    return f"{sum(parsed):.2f}".replace(".", ",")


# ============================================================
# Detecção / estilização de negativos (tksheet)
# ============================================================


def eh_negativo_str(s: Any) -> bool:
    """
    Detecta '-1,23' ou '-01:30' e também o sinal unicode '-'.
    """
    if not isinstance(s, str):
        return False
    st = s.strip()
    if not st:
        return False
    return st.startswith("-") or st.startswith("−")


# ============================================================
# “Compat shims” — para reduzir mudanças nos módulos existentes
# ============================================================


# Mantém o nome usado em extrator/leitor_ponto.py
def tempo_para_decimal(valor: Any) -> float:
    """Compat: alias para hhmm_para_decimal/parse."""
    return hhmm_para_decimal(valor)


# Usado indiretamente em exportações;
# prefira somar_preservando_formato, mas deixamos exposto.
def is_hhmm_string(s: Any) -> bool:
    return _eh_hhmm(s)


def minutes_to_hhmm(total_min: int) -> str:
    return _minutos_para_hhmm(total_min)


# --- helpers extras para UI (tksheet) ---------------------------------
def _to_float_loose(val) -> float | None:
    if isinstance(val, (int, float)):
        return float(val)
    if not isinstance(val, str):
        return None
    s = val.strip()
    if not s:
        return None
    try:
        return float(s.replace(".", "").replace(",", ".", 1))
    except Exception:
        return None


# ============================================================
# Datas: dia da semana em PT-BR e formatação "dd/mm/aa - Seg"
# ============================================================
_DOW_PT = ["Seg", "Ter", "Qua", "Qui", "Sex", "Sáb", "Dom"]


def dia_semana_pt(d) -> str:
    """Retorna abreviação PT-BR do dia da semana (Seg..Dom) para date/datetime."""
    try:
        wd = (d if isinstance(d, date) else datetime.fromisoformat(str(d))).weekday()
    except Exception:
        try:
            # último recurso: converte via pandas se existir
            import pandas as _pd

            wd = _pd.Timestamp(d).weekday()
        except Exception:
            wd = 0
    return _DOW_PT[int(wd) % 7]


def formatar_data_com_dow(
    d, *, estilo: str = "curto", uppercase: bool = True, formato: str = "%d/%m/%y"
) -> str:
    """
    Retorna 'dd/mm/aa - DOW' (curto|longo, maiúsculas opcional).
    """
    try:
        from datetime import date as _d
        from datetime import datetime as _dt

        if isinstance(d, (_d, _dt)):
            dt = d
        else:
            import pandas as _pd

            dt = _pd.Timestamp(d).to_pydatetime()
    except Exception:
        from datetime import datetime as _dt

        dt = _dt.now()

    nomes_curto = ["SEG", "TER", "QUA", "QUI", "SEX", "SÁB", "DOM"]
    nomes_longo = ["segunda", "terça", "quarta", "quinta", "sexta", "sábado", "domingo"]
    idx = dt.weekday()
    nome = nomes_curto[idx] if estilo == "curto" else nomes_longo[idx]
    if uppercase:
        nome = nome.upper()
    elif estilo == "curto":
        nome = nome.capitalize()
    return f"{dt.strftime(formato)} - {nome}"


# ============================================================
# Decimais: formatação com vírgula e zero opcional como '-'
# ============================================================
def formatar_decimal(
    valor, *, casas: int = 2, separador: str = ",", zero_vira_traco: bool = False
) -> str:
    """
    Formata número com 'casas' e separador (padrão ',').
    Se zero_vira_traco=True e |valor| < 1e-12 -> '-'.
    """
    try:
        v = (
            float(str(valor).replace(".", "").replace(",", ".", 1))
            if isinstance(valor, str)
            else float(valor)
        )
    except Exception:
        return str(valor)
    if zero_vira_traco and abs(v) < 1e-12:
        return "-"
    s = f"{v:.{casas}f}"
    return s.replace(".", separador)


# ============================================================
# Formatação de data em pt-BR
# ============================================================
def formatar_data_ptbr(
    dt,
    *,
    formato_data: str = "%d/%m/%y",
    estilo_semana: str = "curto",  # "curto" | "longo"
    uppercase: bool = True,
    separador: str = " - ",
) -> str:
    """
    Ex.: '14/08/25 - QUI' (curto) ou '14/08/25 - QUINTA' (longo).
    Aceita date/datetime/str (melhor-esforço).
    """
    import datetime as _dt

    d = None
    # já é date/datetime?
    if hasattr(dt, "date"):
        try:
            d = dt.date()
        except Exception:
            pass
    if d is None:
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y"):
            try:
                d = _dt.datetime.strptime(str(dt), fmt).date()
                break
            except Exception:
                continue
    if d is None:
        d = _dt.date.today()

    # 0=segunda .. 6=domingo
    nomes_curto = ["SEG", "TER", "QUA", "QUI", "SEX", "SÁB", "DOM"]
    nomes_longo = ["segunda", "terça", "quarta", "quinta", "sexta", "sábado", "domingo"]
    idx = d.weekday()

    if estilo_semana == "longo":
        nome = nomes_longo[idx]
        nome = nome.upper() if uppercase else nome
    else:
        nome = nomes_curto[idx]
        if not uppercase:
            nome = nome.capitalize()

    return f"{d.strftime(formato_data)}{separador}{nome}"


# ============================================================
# Sincronizar rolagem vertical entre múltiplas sheets
# ============================================================
def sincronizar_rolagem_sheets(sheets) -> None:
    """
    Sincroniza a rolagem vertical entre várias tksheets.
    Usa get_visible_rows()/see(...) quando disponível.
    """
    sheets = [s for s in sheets if s is not None]
    if len(sheets) < 2:
        return

    def _top_visible(sheet):
        try:
            r1, r2 = sheet.get_visible_rows()
            return r1
        except Exception:
            return 0

    def _sync_from(source):
        base = _top_visible(source)
        for other in sheets:
            if other is source:
                continue
            try:
                other.see(base, 0)
            except Exception:
                pass

    for s in sheets:
        # evita bind duplicado
        if getattr(s, "_scroll_sync_bound", False):
            continue
        try:
            s.bind("<MouseWheel>", lambda e, src=s: (_sync_from(src), None))
            s.bind("<Button-4>", lambda e, src=s: (_sync_from(src), None))  # Linux up
            s.bind("<Button-5>", lambda e, src=s: (_sync_from(src), None))  # Linux down
            s._scroll_sync_bound = True
        except Exception:
            pass


####################################################
#
#
#####    nova atualização   ##################
def _safe_get_data(sheet):
    try:
        return sheet.get_sheet_data(return_copy=True)  # v6+
    except TypeError:
        try:
            return sheet.get_sheet_data()  # mais antigas
        except Exception:
            try:
                return [list(r) for r in sheet.MT.data]  # último recurso
            except Exception:
                return []


def alinhar_direita_sheet(sheet, *, start_col: int = 0) -> None:
    if sheet is None:
        return
    data = _safe_get_data(sheet)
    cols = len(data[0]) if data else 0

    for c in range(max(0, start_col), cols):
        aligned = False

        # tentativas por coluna (várias APIs de versões diferentes)
        for meth, kwargs in (
            ("align_column", {"c": c, "align": "e"}),
            ("align_columns", {"columns": [c], "align": "e"}),
            ("set_all_cell_alignments", {"c": c, "align": "e"}),
        ):
            try:
                getattr(sheet, meth)(**kwargs)
                aligned = True
                break
            except Exception:
                pass

        # fallback "força bruta": célula a célula
        if not aligned:
            rows = len(data)
            for r in range(rows):
                ok_cell = False
                for meth, kwargs in (
                    ("align_cell", {"r": r, "c": c, "align": "e"}),
                    ("set_cell_alignments", {"r": r, "c": c, "align": "e"}),
                ):
                    try:
                        getattr(sheet, meth)(**kwargs)
                        ok_cell = True
                        break
                    except Exception:
                        pass
                if not ok_cell:
                    # último recurso: ignora
                    continue


def destacar_barra_totais(
    sheet,
    row_index: int | None = None,
    *,
    bg: str = "#000000",
    fg: str = "#ffffff",  # mantemos a assinatura, mas não vamos usar fg aqui
    height: int = 24,
) -> None:
    if sheet is None:
        return
    data = _safe_get_data(sheet)
    r = (len(data) - 1) if row_index is None else row_index
    if r is None or r < 0:
        return

    try:
        # 🔧 Importante: NÃO definir fg aqui — só o bg
        sheet.highlight_rows(r, r, bg=bg)
        try:
            sheet.set_row_height(r, height=height)
        except Exception:
            pass
        return
    except Exception:
        pass

    # Fallback: pinta célula a célula apenas o fundo (sem fg)
    cols = len(data[0]) if data else 0
    for c in range(cols):
        try:
            sheet.highlight_cells(row=r, column=c, bg=bg)
        except Exception:
            pass
    try:
        sheet.set_row_height(r, height=height)
    except Exception:
        pass


def estilizar_negativos_sheet(
    sheet, start_row: int = 0, start_col: int = 0, skip_rows=None
) -> None:
    """
    Pinta em vermelho apenas células cujo TEXTO começa com '-' (ou '−').
    Linhas em skip_rows são ignoradas (ex.: a linha de TOTAIS).
    """
    if sheet is None:
        return

    skip = set(skip_rows or [])
    data = _safe_get_data(sheet)
    for r in range(start_row, len(data)):
        if r in skip:
            continue
        row = data[r]
        for c in range(start_col, len(row)):
            s = "" if row[c] is None else str(row[c]).strip()
            if s[:1] in ("-", "−"):  # só pinta se começa com sinal de menos
                try:
                    sheet.highlight_cells(
                        row=r, column=c, fg="#b00000", font=("Segoe UI", 10, "bold")
                    )
                except Exception:
                    try:
                        sheet.highlight_cells(row=r, column=c, fg="#b00000")
                    except Exception:
                        pass


def ocultar_indices_sheet(sheet):
    """Oculta a coluna de índices (1, 2, 3...) em qualquer versão do tksheet."""
    if sheet is None:
        return
    try:
        # Versões mais novas (>=7.x)
        sheet.show_index = False
        try:
            sheet.row_index_width(0)
        except Exception:
            pass
    except Exception:
        # fallback para versões mais antigas
        try:
            sheet.set_options(show_row_index=False)
            try:
                sheet.set_row_index_width(0)
            except Exception:
                pass
        except Exception:
            pass

# --- helpers novos/ajustados -------------------------------------------------

def ocultar_colunas(sheet, indices):
    """
    Oculta colunas pela largura (width=0) sem afetar cabeçalho/dados.
    Reaplicável em atualizações sucessivas sem 'desocultar' depois.
    """
    if sheet is None:
        return
    if not isinstance(indices, (list, tuple, set)):
        indices = [indices]
    for c in sorted(set(int(i) for i in indices if i is not None and i >= 0)):
        for meth, kwargs in (
            ("set_column_width", {"c": c, "width": 0}),
            ("column_width", {"column": c, "width": 0}),
        ):
            try:
                getattr(sheet, meth)(**kwargs)
                break
            except Exception:
                continue
    # força redesenho sem recalcular largura destas colunas
    for m in ("refresh", "redraw"):
        try:
            getattr(sheet, m)()
            break
        except Exception:
            continue


def finalizar_conjunto_comparacao(
    sheet_boletim,
    sheet_ponto,
    sheet_diferenca,
    *,
    start_col_b: int = 1,
    start_col_p: int = 0,
    start_col_d: int = 0,
    destacar_totais: bool = True,
) -> None:
    """
    Acabamento das três sheets da aba Comparação.
    - Boletim: mantém Data e Boletim visíveis.
    - Ponto/Diferença: OCULTA somente a 1ª coluna (Data = índice 0).
    """
    try:
        finalizar_sheet(
            sheet_boletim,
            start_col=start_col_b,
            total_row=("auto" if destacar_totais else None),
            header_bg="#BAD8F3",
        )
    except Exception:
        pass
    try:
        finalizar_sheet(
            sheet_ponto,
            start_col=start_col_p,
            total_row=("auto" if destacar_totais else None),
            header_bg="#ABE08E",
        )
    except Exception:
        pass
    try:
        finalizar_sheet(
            sheet_diferenca,
            start_col=start_col_d,
            total_row=("auto" if destacar_totais else None),
            header_bg="#F3C1B9",
        )
    except Exception:
        pass

    # 🔒 Regra da aba COMPARAÇÃO:
    #    Boletim → mostra Data/Boletim; Ponto e Diferença → ocultam a coluna 0 (Data).
    try:
        ocultar_colunas(sheet_ponto, [0])
    except Exception:
        pass
    try:
        ocultar_colunas(sheet_diferenca, [0])
    except Exception:
        pass

    # Sincroniza rolagem (opcional)
    try:
        sincronizar_rolagem_sheets([sheet_boletim, sheet_ponto, sheet_diferenca])
    except Exception:
        pass


def finalizar_sheet(
    sheet,
    *,
    start_col: int = 0,
    total_row="auto",
    header_bg: str = "#e9edf5",
) -> None:
    """
    Acabamento de uma sheet:
    - cabeçalho (bg, alinhamento),
    - alinhamento numérico à direita a partir de start_col,
    - barra preta na linha de totais (com texto branco/vermelho),
    - ajuste de larguras automáticas, **preservando** colunas já ocultas (width=0).
    """
    if sheet is None:
        return

    # Cabeçalho
    try:
        sheet.headers_align = "center"
    except Exception:
        pass
    try:
        sheet.headers_wrap = True
    except Exception:
        pass
    try:
        sheet.headers_bg(header_bg)
    except Exception:
        try:
            sheet.set_options(header_bg=header_bg)
        except Exception:
            pass

    alinhar_direita_sheet(sheet, start_col=start_col)
    ocultar_indices_sheet(sheet)

    MIN_PX = 75  # ≈ 10 caracteres

    data = _safe_get_data(sheet)
    r_total = (len(data) - 1) if (total_row == "auto" and data) else total_row

    # 0) Escrever TOTAIS (antes da barra preta)
    if isinstance(r_total, int) and r_total >= 0 and data:
        rows = len(data)
        cols = len(data[0]) if data else 0

        # rótulo "TOTAIS:" na coluna logo antes da primeira numérica
        try:
            rotulo_col = max(0, start_col - 1)
            sheet.set_cell_data(r_total, rotulo_col, "TOTAIS:")
        except Exception:
            pass

        # soma coluna a coluna, preservando formato (HH:MM vs decimal)
        for c in range(max(0, start_col), cols):
            try:
                col_vals = [data[r][c] if r != r_total else "" for r in range(rows)]
            except Exception:
                col_vals = []
            total_txt = somar_preservando_formato(col_vals, default_decimal=True)
            # default_decimal=True => usa vírgula e 2 casas quando decimal
            try:
                sheet.set_cell_data(r_total, c, total_txt)
            except Exception:
                pass

        # refaz cache de dados para as etapas seguintes (pintura/ajuste)
        data = _safe_get_data(sheet)

    # 1) barra preta na linha de total
    if isinstance(r_total, int) and r_total >= 0:
        destacar_barra_totais(sheet, r_total)
        for m in ("refresh", "redraw"):
            try:
                getattr(sheet, m)()
                break
            except Exception:
                continue

    # 2) negativos (pula a linha total)
    skip = {r_total} if isinstance(r_total, int) and r_total >= 0 else None
    estilizar_negativos_sheet(sheet, start_row=0, start_col=start_col, skip_rows=skip)

    # 3) linha TOTAL: texto branco e vermelho nos negativos (mantendo BG preto)
    if isinstance(r_total, int) and r_total >= 0:
        try:
            try:
                total_cols = sheet.get_total_columns()
                if not total_cols:
                    raise Exception()
            except Exception:
                total_cols = len(data[0]) if data else 0

            for c in range(max(0, start_col), total_cols):
                try:
                    s = str(data[r_total][c]).strip()
                except Exception:
                    s = ""
                eh_neg = s[:1] in ("-", "−")
                cor_txt = "#ff3b30" if eh_neg else "#ffffff"
                ok = False
                for meth, kwargs in (
                    ("highlight_cells",
                     {"row": r_total, "column": c, "bg": "#000000", "fg": cor_txt,
                      "font": ("Segoe UI", 10, "bold")}),
                    ("highlight_cells",
                     {"row": r_total, "column": c, "bg": "#000000", "fg": cor_txt}),
                ):
                    try:
                        getattr(sheet, meth)(**kwargs)
                        ok = True
                        break
                    except Exception:
                        continue
                if not ok:
                    try:
                        sheet.highlight_cells(row=r_total, column=c, fg=cor_txt)
                        sheet.highlight_cells(row=r_total, column=c, bg="#000000")
                    except Exception:
                        pass
        except Exception:
            pass

    # 4) Ajuste de larguras — **NÃO** reabrir colunas já ocultas (width==0)
    try:
        total_cols = sheet.get_total_columns()
    except Exception:
        data = sheet.get_sheet_data(return_copy=True)
        total_cols = len(data[0]) if data else 0

    for c in range(total_cols):
        # se já está oculta, não recalcular largura
        try:
            w_atual = sheet.get_column_width(c)
            if w_atual == 0:
                continue
        except Exception:
            pass

        try:
            # auto com mínimo
            sheet.set_column_width(c, width="auto")
            largura_auto = sheet.get_column_width(c)
            largura_final = max(MIN_PX, largura_auto)
            sheet.set_column_width(c, width=largura_final)
        except Exception:
            try:
                sheet.column_width(column=c, width=75)
            except Exception:
                pass

    for m in ("refresh", "redraw", "Recalculate"):
        try:
            getattr(sheet, m)()
            break
        except Exception:
            continue

# ============================================================
# Excel (openpyxl) – helpers de cabeçalho e formatação
# ============================================================


STARTROW_DADOS_XLSX = 7  # pandas startrow=7 -> linha 8 no Excel (1-based)

def aplicar_cabecalho_excel(ws, titulo, contrato, dt_ini, dt_fim, boletins, df_cols):
    """
    Escreve título + metadados padronizados no topo da aba do Excel e congela o cabeçalho.
    - ws: Worksheet do openpyxl
    - titulo: str
    - contrato: str|int|lista (é convertido para string)
    - dt_ini, dt_fim: datetime
    - boletins: texto já pronto (ex.: '123, 456, 789' ou '-')
    - df_cols: lista de nomes de colunas (para saber largura/mesclas)
    """
    max_col = max(1, len(df_cols))

    # Linha 1: Título centralizado e em negrito
    ws.append([str(titulo)])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Linhas 3..5: metadados
    def _bold(r, c, txt):
        cell = ws.cell(row=r, column=c, value=txt)
        cell.font = Font(bold=True)
        return cell

    _bold(3, 1, "Contrato:")
    ws.cell(row=3, column=2, value="" if contrato is None else str(contrato))

    _bold(4, 1, "Período de Apuração:")
    try:
        periodo = f"{dt_ini:%d/%m/%Y} a {dt_fim:%d/%m/%Y}"
    except Exception:
        periodo = ""
    ws.cell(row=4, column=2, value=periodo)

    _bold(5, 1, "Boletins Incluídos:")
    ws.cell(row=5, column=2, value=str(boletins) if boletins is not None else "-")

    # Congela antes da primeira linha de dados
    ws.freeze_panes = f"A{STARTROW_DADOS_XLSX + 1}"

# utils/formatadores.py (trecho)
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

THIN = Side(style="thin", color="000000")
BORDER_THIN = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

HEADER_FILL = PatternFill("solid", fgColor="F2F2F2")
HEADER_FONT = Font(bold=True)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")

# Em pt-BR o Excel interpreta esse formato com vírgula decimal
NUM_FMT_2C = "#.##0,00"
NUM_FMT_INT = "#.##0"

def _is_number(value) -> bool:
    try:
        return (value is not None) and (str(value).strip() != "") and (float(str(value).replace(",", ".") ) or True)
    except Exception:
        return False

def _infer_number_format(value):
    """Tenta escolher formato inteiro vs 2 casas."""
    try:
        v = float(str(value).replace(",", "."))
        return NUM_FMT_INT if abs(v) == int(abs(v)) else NUM_FMT_2C
    except Exception:
        return None

def _auto_width_for_column(ws: Worksheet, col_idx: int, start_row: int, end_row: int, min_w=12, max_w=42) -> int:
    """Calcula largura ideal com base no conteúdo visível (sem estilos)."""
    letter = get_column_letter(col_idx)
    max_len = 0

    # inclui o header na conta
    hdr_cell = ws.cell(row=start_row, column=col_idx)
    if hdr_cell.value is not None:
        max_len = max(max_len, len(str(hdr_cell.value)))

    for r in range(start_row + 1, end_row + 1):
        cell = ws.cell(row=r, column=col_idx)
        val = cell.value
        if val is None:
            continue
        s = str(val)
        # números costumam ficar um pouco mais largos por alinhamento à direita
        if _is_number(val):
            s = f"{val}"
        # limitações básicas para não estourar
        max_len = max(max_len, len(s))

    # ajuste heurístico (fonte padrão Calibri 11 ~ 0.9 char/px)
    width = min(max(min_w, int(max_len * 1.1) + 2), max_w)
    ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, width)
    return width

def formatar_planilha_excel(
    ws: Worksheet,
    startrow: int = 1,
    header_row_height: int = 22,
    zebra: bool = True,
    aplicar_bordas: bool = True,
    congelar_cabecalho: bool = True,
    aplicar_autofiltro: bool = True,
    detectar_numeros: bool = True,
):
    """
    Aplica formatação padrão na planilha ativa.

    Parâmetros:
      - ws: Worksheet (openpyxl)
      - startrow: linha do cabeçalho dos dados (1-based)
    Requisitos:
      - Cabeçalho começa em 'startrow'
      - Dados (se existirem) começam em 'startrow + 1'
    """

    if ws.max_row < startrow:
        return  # nada para formatar

    first_data_row = startrow + 1
    last_row = ws.max_row
    last_col = ws.max_column

    # 1) Cabeçalho (negrito, fill cinza claro, centralizado, altura)
    ws.row_dimensions[startrow].height = header_row_height
    for c in range(1, last_col + 1):
        cell = ws.cell(row=startrow, column=c)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        if aplicar_bordas:
            cell.border = BORDER_THIN

    # 2) Linhas de dados: alinhamento e zebra
    if last_row >= first_data_row:
        for r in range(first_data_row, last_row + 1):
            is_even = (r - first_data_row) % 2 == 0
            fill = PatternFill("solid", fgColor="FFFFFF")
            if zebra and is_even:
                fill = PatternFill("solid", fgColor="FAFAFA")

            for c in range(1, last_col + 1):
                cell = ws.cell(row=r, column=c)

                # alinhamento padrão
                if _is_number(cell.value):
                    cell.alignment = RIGHT
                else:
                    cell.alignment = LEFT

                # bordas finas
                if aplicar_bordas:
                    cell.border = BORDER_THIN

                # zebra
                cell.fill = fill

                # formatação numérica (pt-BR: vírgula decimal)
                if detectar_numeros and _is_number(cell.value):
                    fmt = _infer_number_format(cell.value) or NUM_FMT_2C
                    cell.number_format = fmt

    # 3) Ajuste de largura de colunas (mín. 12, máx. 42)
    for c in range(1, last_col + 1):
        _auto_width_for_column(ws, c, startrow, last_row, min_w=12, max_w=42)

    # 4) Freeze panes (congela linha do cabeçalho)
    if congelar_cabecalho:
        # congela logo abaixo do cabeçalho, na primeira coluna
        ws.freeze_panes = ws.cell(row=first_data_row, column=1).coordinate

    # 5) AutoFilter sobre o range de dados (inclui cabeçalho)
    if aplicar_autofiltro:
        ws.auto_filter.ref = f"{ws.cell(startrow, 1).coordinate}:{ws.cell(last_row, last_col).coordinate}"



@dataclass
class ColFmt:
    # nome da coluna como aparece no DataFrame
    name: str
    number_format: Optional[str] = None   # e.g. NUM_FMT_2C, NUM_FMT_MONEY
    align: Optional[str] = None           # "left" | "right" | "center"
    width: Optional[int] = None

def _is_number(value) -> bool:
    return isinstance(value, (int, float)) and value is not None

def _choose_align(fmt: Optional[str]) -> Alignment:
    if fmt is None:
        return LEFT
    return RIGHT

def _auto_width(ws: Worksheet, start_row: int, end_row: int, min_w=12, max_w=42):
    for c in range(1, ws.max_column + 1):
        letter = get_column_letter(c)
        max_len = 0
        # header
        val = ws.cell(row=start_row, column=c).value
        if val is not None:
            max_len = max(max_len, len(str(val)))
        # data
        for r in range(start_row + 1, end_row + 1):
            v = ws.cell(row=r, column=c).value
            if v is None:
                continue
            s = str(v)
            max_len = max(max_len, len(s))
        width = min(max(min_w, int(max_len * 1.1) + 2), max_w)
        ws.column_dimensions[letter].width = max(ws.column_dimensions[letter].width or 0, width)

def escrever_df_formatado(
    wb: Workbook,
    sheet_name: str,
    df: pd.DataFrame,
    *,
    startrow: int = 1,
    startcol: int = 1,
    col_formats: Optional[Sequence[ColFmt]] = None,
    zebra: bool = True,
    freeze_header: bool = True,
    autofilter: bool = True,
    destacar_total: bool = True,
    aplicar_bordas: bool = True,
    cond_format_diferencas: Optional[str] = None,  # nome da coluna que recebe vermelho/verde
) -> Worksheet:
    """
    Escreve o DataFrame e aplica formatação profissional imediatamente.
    Retorna a Worksheet (para adicionar gráficos, etc., se quiser).
    """
    ws = wb.create_sheet(title=sheet_name)

    # 1) Cabeçalho
    for j, col in enumerate(df.columns, start=startcol):
        c = ws.cell(row=startrow, column=j, value=col)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        if aplicar_bordas:
            c.border = BORDER_THIN

    # 2) Dados
    for i, (_, row) in enumerate(df.iterrows(), start=startrow + 1):
        is_even = (i - (startrow + 1)) % 2 == 0
        fill = PatternFill("solid", fgColor=("FAFAFA" if (zebra and is_even) else "FFFFFF"))
        for j, col in enumerate(df.columns, start=startcol):
            val = row[col]
            cell = ws.cell(row=i, column=j, value=val)
            # alinhamento padrão
            cell.alignment = RIGHT if _is_number(val) else LEFT
            # bordas e zebra
            if aplicar_bordas:
                cell.border = BORDER_THIN
            cell.fill = fill

    last_row = startrow + len(df)
    last_col = startcol + len(df.columns) - 1

    # 3) Formatação por coluna (número/alinhamento/largura)
    if col_formats:
        name_to_idx = {name: idx for idx, name in enumerate(df.columns, start=startcol)}
        for cf in col_formats:
            if cf.name not in name_to_idx:
                continue
            j = name_to_idx[cf.name]
            # number_format
            if cf.number_format:
                for i in range(startrow + 1, last_row + 1):
                    ws.cell(row=i, column=j).number_format = cf.number_format
            # align
            if cf.align:
                align = {"left": LEFT, "right": RIGHT, "center": CENTER}.get(cf.align, None)
                if align:
                    for i in range(startrow + 1, last_row + 1):
                        ws.cell(row=i, column=j).alignment = align
            # width
            if cf.width:
                ws.column_dimensions[get_column_letter(j)].width = cf.width

    # 4) Cond. formatting (diferenças: vermelho negativo, verde positivo)
    if cond_format_diferencas and cond_format_diferencas in df.columns:
        j = list(df.columns).index(cond_format_diferencas) + startcol
        col_letter = get_column_letter(j)
        data_range = f"{col_letter}{startrow+1}:{col_letter}{last_row}"
        ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator='lessThan', formula=['0'], stopIfTrue=False,
                       font=Font(color="9C0006"), fill=PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"))
        )
        ws.conditional_formatting.add(
            data_range,
            CellIsRule(operator='greaterThan', formula=['0'], stopIfTrue=False,
                       font=Font(color="006100"), fill=PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"))
        )

    # 5) Destacar linha TOTAL (busca na 1ª coluna por 'TOTAL')
    if destacar_total and len(df) > 0:
        total_rows = []
        primeira_coluna = startcol
        for i in range(startrow + 1, last_row + 1):
            v = ws.cell(row=i, column=primeira_coluna).value
            if isinstance(v, str) and v.strip().upper() == "TOTAL":
                total_rows.append(i)
        if total_rows:
            fill_total = PatternFill("solid", fgColor="D9D9D9")
            font_total = Font(bold=True)
            for i in total_rows:
                for j in range(startcol, last_col + 1):
                    cell = ws.cell(row=i, column=j)
                    cell.fill = fill_total
                    cell.font = font_total
                    if aplicar_bordas:
                        cell.border = BORDER_THIN

    # 6) AutoFilter + Freeze
    if autofilter:
        ws.auto_filter.ref = f"{ws.cell(startrow, startcol).coordinate}:{ws.cell(last_row, last_col).coordinate}"
    if freeze_header:
        ws.freeze_panes = ws.cell(row=startrow + 1, column=startcol + 1).coordinate

    # 7) Auto largura (mantém width manual se já definido)
    _auto_width(ws, startrow, last_row, min_w=12, max_w=42)
    return ws
